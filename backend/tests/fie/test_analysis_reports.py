"""analysis_reports: precise parser for the PSX year-{year}.xlsx fundamentals file.

The sheet has a 4-row preamble, a header row (Symbol / Name of Company / …), and
sector section-header rows (no Sr. No.) carried down over the company rows. Values are
Rs. million (financials), million (shares), and % of face value (dividends). One typed
record per company.
"""

import io
import os

import pytest
from openpyxl import Workbook

from app.engines.fie.apis import parsers as P
from app.engines.fie.apis.registry import BY_NAME

_HEADER = ["Sr. No.", "Symbol", "Name of Company", "Year End", "Paid up Capital",
           "Face Value", "Number of Shares", "Shareholders' Equity", "Total Assets",
           "Sales / Total Income", "Financial Charges", "Profit Before Taxation",
           "Taxation", "Profit After Taxation", "Cash Dividend", "Stock Dividend",
           "Total Dividend", "Right Issue", "Number of Shareholders"]


def _make_xlsx() -> bytes:
    """Build a tiny workbook mirroring the real layout (preamble, header at row 2,
    sector section headers, company rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Data Entry"
    ws["D1"] = "DATA FOR THE YEAR 2025"                       # banner
    for c, label in enumerate(_HEADER, start=2):              # header on row 2, cols B..T
        ws.cell(row=2, column=c, value=label)
    ws.cell(row=4, column=6, value="(Rs. in million)")        # units row (ignored)
    # sector header (no Sr. No.; ordinal in Symbol col, sector name in Name col)
    ws.cell(row=6, column=3, value=1)
    ws.cell(row=6, column=4, value="AUTOMOBILE ASSEMBLER")
    # MTL company row (Sr.No in col B): values keyed to real file
    mtl = [1, "MTL", "Millat Tractors Limited", "2025-06-30", 1995.16, 10, 199.516,
           8076.3, 32988.59, 52108.997, 2172.644, 8039.66, 1666.732, 6372.928,
           600, 0, 600, None, 15461]
    for c, v in enumerate(mtl, start=2):
        ws.cell(row=7, column=c, value=v)
    # second sector + company
    ws.cell(row=8, column=3, value=2)
    ws.cell(row=8, column=4, value="CEMENT")
    luck = [1, "LUCK", "Lucky Cement Limited", "2025-06-30", 2930, 2, 1465,
            175910.4, 266748.03, 124511.744, 1370.569, 46993.459, 13901.297, 33092.162,
            200, 0, 200, None, 23877]
    for c, v in enumerate(luck, start=2):
        ws.cell(row=9, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_company_fundamentals_no_column_drift():
    recs = P.parse_analysis_report_xlsx(_make_xlsx())
    assert len(recs) == 2                                     # 2 companies, sector rows excluded
    mtl = {r["symbol"]: r for r in recs}["MTL"]
    assert mtl["name"] == "Millat Tractors Limited"
    assert mtl["sector"] == "AUTOMOBILE ASSEMBLER" and mtl["fiscal_year"] == 2025
    assert mtl["year_end"] == "2025-06-30"
    # equity / assets / sales kept distinct (the easy place to drift)
    assert mtl["equity"] == 8076.3
    assert mtl["total_assets"] == 32988.59
    assert mtl["sales"] == 52108.997
    assert mtl["pat"] == 6372.928 and mtl["pbt"] == 8039.66
    assert mtl["shares_m"] == 199.516 and mtl["shareholders"] == 15461
    assert mtl["cash_dividend_pct"] == 600 and mtl["right_issue_pct"] is None


def test_sector_carried_down():
    by = {r["symbol"]: r for r in P.parse_analysis_report_xlsx(_make_xlsx())}
    assert by["LUCK"]["sector"] == "CEMENT"          # picked up the 2nd section header
    assert by["LUCK"]["equity"] == 175910.4


def test_malformed_safe():
    assert P.parse_analysis_report_xlsx(b"not an xlsx") == []
    assert P.parse_analysis_report_xlsx(123) == []
    assert P.parse_analysis_report_xlsx("notes.txt") == []   # non-xlsx path ignored


def test_registry_entry():
    a = BY_NAME["analysis_reports"]
    assert a.method == "GET" and a.response_type == "xlsx"
    assert "{year}" in a.endpoint and a.dynamic_params == ("year",)
    assert a.parser_fn is P.parse_analysis_report_xlsx


# Optional: validate against the real downloaded file if it's present locally.
_REAL = r"C:\Users\ibrahim.ijaz\Downloads\year-2025.xlsx"


@pytest.mark.skipif(not os.path.exists(_REAL), reason="real year-2025.xlsx not present")
def test_real_file_sample():
    recs = P.parse_analysis_report_xlsx(_REAL)
    assert len(recs) > 400                                   # ~535 companies
    by = {r["symbol"]: r for r in recs}
    assert by["LUCK"]["sector"] == "CEMENT" and by["LUCK"]["pat"] == 33092.162
    assert by["786"]["symbol"] == "786"                      # numeric ticker stays a company
