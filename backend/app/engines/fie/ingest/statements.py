"""Parse statement (headline) and detail sheets into long-format records.

Per-sheet header/band detection is mandatory: headline ``P&L`` has its year
header on row 4 starting at column C, while detail ``PL1`` has it on row 3
starting at column B (docs/fie_phase0_foundation.md §2.2, §2.3).
"""

from __future__ import annotations

import re
from typing import Any, Optional

from openpyxl.utils import get_column_letter

from ..ontology import MetricOntology
from .classify import statement_of

_YEAR_MIN, _YEAR_MAX = 1990, 2100
_NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")


def _coerce_number(v: Any) -> Optional[float]:
    """Coerce a cell value to float, handling '(1,234)', '–', '', None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in {"", "-", "–", "—", "N/A", "n/a", "nil", "Nil"}:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "").replace("–", "-").replace("—", "-")
    if not _NUM_RE.match(s):
        return None
    val = float(s)
    return -val if neg else val


def _is_year(v: Any) -> bool:
    return isinstance(v, int) and _YEAR_MIN <= v <= _YEAR_MAX


def detect_header_row(ws, scan_rows: int = 8) -> Optional[int]:
    """Row index (1-based) whose cells are mostly fiscal-year integers."""
    best_row, best_count = None, 0
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        count = sum(1 for c in range(1, ws.max_column + 1) if _is_year(ws.cell(r, c).value))
        if count > best_count:
            best_row, best_count = r, count
    return best_row if best_count >= 2 else None


def _column_year_map(ws, header_row: int) -> dict[int, int]:
    return {
        c: ws.cell(header_row, c).value
        for c in range(1, ws.max_column + 1)
        if _is_year(ws.cell(header_row, c).value)
    }


def _column_period_map(ws, header_row: int, col_years: dict[int, int]) -> dict[int, str]:
    """Spread Historical/Forecasted band labels (row above header) across columns."""
    band_row = header_row - 1
    periods: dict[int, str] = {}
    current = "historical"
    if band_row >= 1:
        # walk columns left->right, carrying the most recent band label forward
        for c in range(1, ws.max_column + 1):
            label = ws.cell(band_row, c).value
            if label:
                lab = str(label).strip().lower()
                if "forecast" in lab:
                    current = "forecasted"
                elif "histor" in lab:
                    current = "historical"
            if c in col_years:
                periods[c] = current
    if not periods:  # no band row -> assume all historical
        periods = {c: "historical" for c in col_years}
    return periods


def _looks_like_section(label: str, has_values: bool) -> bool:
    """All-caps-ish header rows with no values are section separators."""
    if has_values:
        return False
    s = (label or "").strip()
    if not s:
        return False
    letters = [ch for ch in s if ch.isalpha()]
    return bool(letters) and all(ch.isupper() for ch in letters)


def parse_grid_sheet(ws, *, level: str, ontology: MetricOntology,
                     value_getter=None) -> list[dict]:
    """Parse a statement or detail grid into long-format value records.

    Returns one record per (data row, year column) with a non-skipped label.
    ``value_getter(coord)`` (optional) supplies effective values for cells whose
    cached value is None (uncalculated formulas); falls back to the cached value.
    """
    header_row = detect_header_row(ws)
    if header_row is None:
        return []
    col_years = _column_year_map(ws, header_row)
    col_periods = _column_period_map(ws, header_row, col_years)
    sheet_title = ws.title
    statement = statement_of(sheet_title)

    # find the note column (header cell == "Notes") if present
    note_col = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value or "").strip().lower() == "notes":
            note_col = c
            break

    records: list[dict] = []
    section: Optional[str] = None

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        label = str(label).strip() if label is not None else ""
        row_values = {}
        for c in col_years:
            coord = f"{get_column_letter(c)}{r}"
            v = _coerce_number(ws.cell(r, c).value)
            if v is None and value_getter is not None:
                v = value_getter(coord)
            row_values[c] = v
        has_values = any(v is not None for v in row_values.values())

        if not label and not has_values:
            continue
        if _looks_like_section(label, has_values):
            section = label.strip()
            continue
        if not label:
            continue

        metric = ontology.canonical(label, sheet=sheet_title)
        note_ref = None
        if note_col is not None:
            nv = ws.cell(r, note_col).value
            note_ref = str(nv).strip() if nv not in (None, "") else None

        for c, year in col_years.items():
            records.append({
                "statement": statement,
                "level": level,
                "sheet": sheet_title,
                "cell": f"{get_column_letter(c)}{r}",
                "label": label,
                "section": section,
                "metric": metric,
                "note_ref": note_ref,
                "year": int(year),
                "period_type": col_periods.get(c, "historical"),
                "value": row_values[c],
            })
    return records
