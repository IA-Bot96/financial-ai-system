"""Shared workbook styling — matches the Millat/Lucky template look.

Palette extracted from the templates' breakdown sheets:
  navy 1F4E79 (title + header, white text), light blue BDD7EE (unit line),
  pale blue D6E4F0 (section headers), green D5E8D4 (totals).
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill

NAVY = "FF1F4E79"
UNIT_BLUE = "FFBDD7EE"
SECTION_BLUE = "FFD6E4F0"
TOTAL_GREEN = "FFD5E8D4"
WHITE = "FFFFFFFF"
BLACK = "FF000000"

NUMBER_FORMAT = r"#,##0;(#,##0);\-"
CONFIDENCE_FORMAT = "0.00"

# Fonts
TITLE_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FONT = Font(name="Arial", size=9, bold=True, color=WHITE)
UNIT_FONT = Font(name="Arial", size=9, bold=True, color=BLACK)
LABEL_FONT = Font(name="Arial", size=9, bold=False, color=BLACK)
VALUE_FONT = Font(name="Arial", size=9, bold=False, color=BLACK)
TOTAL_FONT = Font(name="Arial", size=9, bold=True, color=BLACK)
SECTION_FONT = Font(name="Arial", size=9, bold=True, color=BLACK)

# Fills
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
UNIT_FILL = PatternFill("solid", fgColor=UNIT_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=SECTION_BLUE)
TOTAL_FILL = PatternFill("solid", fgColor=TOTAL_GREEN)

# Alignments
LEFT = Alignment(horizontal="left")
RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center")

LABEL_COL_WIDTH = 52
VALUE_COL_WIDTH = 18

_TOTAL_PREFIXES = ("total", "gross ", "net ", "operating profit", "profit ", "subtotal")


def is_section_header(label: str) -> bool:
    letters = [c for c in label if c.isalpha()]
    if len(letters) < 3:
        return False
    return sum(c.isupper() for c in letters) / len(letters) >= 0.7


def is_total_row(label: str) -> bool:
    low = label.strip().lower()
    return low.startswith(_TOTAL_PREFIXES)
