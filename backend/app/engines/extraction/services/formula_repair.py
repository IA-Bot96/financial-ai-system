"""Template-author formula repair (Bucket A).

Some hand-authored templates ship with two defects that survive into the filled
workbook because our writer only places *values* and never rewrites the template's
formula cells:

  A1 — Frozen absolute reference.  A subtotal whose formula was dragged across the
       year columns with the column LOCKED (`=$C$10+$C$17`), so every year column
       recomputes the first year's value instead of its own.

  A2 — Hard-coded check cell.  A self-check / difference row (`=C57-C28`) where the
       author replaced the formula with a literal `0` in later year columns, which
       MASKS any imbalance (the row always reads "balanced").

Both are the template author's defect, not our data — verified by opening the blank
template. This pass repairs them generically on the written workbook:

  A1: within a row's year-column band, if >=2 formula cells are byte-identical and
      reference exactly one band column via an absolute ref, re-relativize that
      column so each cell tracks its own column (`$C$10` -> `D10` in column D, ...).
  A2: a literal `0` year-cell whose row has a sibling FORMULA cell -> translate the
      sibling's (relative) formula into the `0` cell (`C57-C28` -> `D57-D28`).

Every rewrite is returned for logging so the repair is auditable. Guardrails keep it
from touching legitimately-fixed references or plain data rows (see inline notes).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Optional

from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger

logger = get_logger(__name__)

# A single A1-style cell reference: optional $col, col letters, optional $row, row.
_REF = re.compile(r"(?<![A-Za-z0-9_'])(\$?)([A-Z]{1,3})(\$?)(\d+)")
_HEADER_SCAN_ROWS = 8


@dataclass
class Repair:
    sheet: str
    coordinate: str
    kind: str            # "frozen-ref" | "constant-check"
    before: object
    after: str

    def __str__(self) -> str:
        return f"{self.sheet}!{self.coordinate} [{self.kind}] {self.before!r} -> {self.after!r}"


def _year_columns(ws) -> tuple[int, list[int]]:
    """(header_row, [year-data column indices]); ([],) when no >=2-year header found.

    Mirrors the validation pass: the first row in the top rows holding >=2 year-like
    integers defines the year-column band."""
    best_cols: list[int] = []
    best_row = 0
    for r in range(1, min(ws.max_row, _HEADER_SCAN_ROWS) + 1):
        cols = [c for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, (int, float))
                and not isinstance(ws.cell(r, c).value, bool)
                and 1990 <= int(ws.cell(r, c).value) <= 2100]
        if len(cols) > len(best_cols):
            best_cols, best_row = cols, r
    return (best_row, best_cols) if len(best_cols) >= 2 else (0, [])


def _band_columns_in(formula: str, band_letters: set[str]) -> set[str]:
    """Band column letters referenced by LOCAL (same-sheet) refs in the formula."""
    if "!" in formula:        # cross-sheet pull — leave alone (these shift correctly)
        return set()
    return {m.group(2) for m in _REF.finditer(formula) if m.group(2) in band_letters}


def _rerelativize(formula: str, anchor_letter: str, target_letter: str) -> str:
    """Replace refs to `anchor_letter` with a RELATIVE ref to `target_letter`,
    keeping the row, and leave every other reference untouched."""
    def sub(m: re.Match) -> str:
        _col_abs, col, _row_abs, row = m.groups()
        if col == anchor_letter:
            return f"{target_letter}{row}"   # relative, tracks this column
        return m.group(0)                    # untouched (other columns / fixed refs)
    return _REF.sub(sub, formula)


def _strip_abs(formula: str) -> str:
    """Drop the `$` lock markers so two formulas can be compared for shape."""
    return formula.replace("$", "")


def _repair_frozen(ws, header_row: int, band: list[int]) -> list[Repair]:
    """A1 — re-relativize column-locked formulas that were drag-copied to the wrong
    column. The defect signature is PER CELL: the cell's only band reference is an
    ABSOLUTE lock on a *different* band column, and the cell is a byte-for-byte (modulo
    `$`) copy of that column's own SELF-referential cell — i.e. a frozen drag. A cell
    that already references its own column, or whose neighbour isn't self-referential
    (an intentional cross-column formula like a growth chain), is left alone."""
    band_letters = {get_column_letter(c) for c in band}
    col_of_letter = {get_column_letter(c): c for c in band}
    out: list[Repair] = []
    for r in range(header_row + 1, ws.max_row + 1):
        for c in band:
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):         # read-only non-anchor of a merge
                continue
            val = cell.value
            if not (isinstance(val, str) and val.startswith("=")):
                continue
            refd = _band_columns_in(val, band_letters)
            if len(refd) != 1:                       # must lock exactly one band column
                continue
            anchor = next(iter(refd))
            own = get_column_letter(c)
            if anchor == own:                        # already tracks its own column — fine
                continue
            if not re.search(rf"\${anchor}\$?\d+", val):   # the lock ($) is the defect
                continue
            # The anchor column's cell in this row must itself be a SELF-referential
            # formula, and `val` must be a copy of it (proves an intended per-column
            # pattern that got frozen) — not a deliberate reference to another column.
            ref_cell = ws.cell(r, col_of_letter[anchor])
            ref_val = ref_cell.value
            if not (isinstance(ref_val, str) and ref_val.startswith("=")):
                continue
            if _band_columns_in(ref_val, band_letters) != {anchor}:
                continue
            if _strip_abs(val) != _strip_abs(ref_val):
                continue
            new = _rerelativize(val, anchor, own)
            if new != val:
                out.append(Repair(ws.title, cell.coordinate, "frozen-ref", val, new))
                cell.value = new
    return out


def _repair_constant_checks(ws, header_row: int, band: list[int]) -> list[Repair]:
    """A2 — restore literal-0 year cells from a sibling formula in the same row."""
    out: list[Repair] = []
    for r in range(header_row + 1, ws.max_row + 1):
        cells = [ws.cell(r, c) for c in band]
        formula_cell = next((cell for cell in cells
                             if isinstance(cell.value, str) and cell.value.startswith("=")
                             and "!" not in cell.value), None)
        if formula_cell is None:
            continue
        for cell in cells:
            # Only literal 0 (the masked-check pattern); never overwrite real data.
            if isinstance(cell, MergedCell) or cell is formula_cell \
                    or not isinstance(cell.value, (int, float)):
                continue
            if isinstance(cell.value, bool) or cell.value != 0:
                continue
            new = Translator(formula_cell.value, origin=formula_cell.coordinate)\
                .translate_formula(cell.coordinate)
            out.append(Repair(ws.title, cell.coordinate, "constant-check", cell.value, new))
            cell.value = new
    return out


def repair_template_formulas(workbook_path, save: bool = True) -> list[Repair]:
    """Repair Bucket-A defects in the written workbook in place. Returns the list of
    repaired cells (for logging / the manifest). Best-effort: a malformed sheet is
    skipped, never fatal."""
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=False)
    repairs: list[Repair] = []
    for ws in wb.worksheets:
        try:
            header_row, band = _year_columns(ws)
            if not band:
                continue
            repairs.extend(_repair_frozen(ws, header_row, band))
            repairs.extend(_repair_constant_checks(ws, header_row, band))
        except Exception as exc:  # noqa: BLE001 — repair is best-effort
            logger.warning("Formula repair skipped sheet %r: %s", ws.title, exc)
    if repairs and save:
        wb.save(workbook_path)
    wb.close()
    return repairs
