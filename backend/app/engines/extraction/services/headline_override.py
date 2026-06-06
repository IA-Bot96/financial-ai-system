"""Bucket B — headline face-truth override.

The output statements (P&L, Balance Sheet) are the deliverable; their breakdown
sheets are supporting notes. When a breakdown's leaves are incomplete or
mis-placed, the output cell that pulls/sums them won't tie out to the audited
TrustedFaceIndex (which we proved matches the PDF). For those HEADLINE cells we
substitute the audited face value directly — with a provenance comment — so the
delivered statements are correct even when the supporting detail is incomplete.

Scope guards keep this honest and narrow:
  - only the declared OUTPUT sheets (never breakdown notes),
  - only rows resolving to a KEY headline metric,
  - only (metric, year) pairs we actually have confident face truth for,
  - only cells that DON'T already tie out — a correct live formula is preserved.
Breakdown-sheet mismatches are intentionally left untouched and stay flagged in
the Validation Ledger as incomplete detail.
"""
from __future__ import annotations

from dataclasses import dataclass

from app.core.logging import get_logger

logger = get_logger(__name__)

# Sign convention by metric (authoritative for these unambiguous metrics; see resolve loop).
# Expenses/deductions are ALWAYS negative on an additive P&L; revenue, balance-sheet figures,
# and "other income" (income by definition) are always positive. Profit SUBTOTALS are
# intentionally OMITTED — they can legitimately be negative (a loss year) — so they infer
# the sign from a sibling formula / keep face truth's own sign.
_ALWAYS_NEGATIVE = frozenset({"cost_of_sales", "finance_cost", "tax_expense", "taxation", "income_tax"})
_ALWAYS_POSITIVE = frozenset({
    "revenue", "total_assets", "total_liabilities", "total_equity_and_liabilities",
    "non_current_assets", "current_assets", "non_current_liabilities", "current_liabilities", "equity",
    "other_income",   # "other income" is income by definition -> never emit negative (#4)
})


@dataclass
class Override:
    sheet: str
    coordinate: str
    metric: str
    year: int
    was: object
    value: float
    source: str
    report_file: object = None     # structured provenance so the Source Ledger row
    report_year: object = None     # carries the same columns as a normal write row
    page: object = None
    table_id: object = None

    def __str__(self) -> str:
        return f"{self.sheet}!{self.coordinate} {self.metric}/{self.year}: {self.was!r} -> {self.value:,.0f}"


def override_headline_metrics(workbook_path, company, tieout, output_sheets,
                              annotate: bool = True) -> list[Override]:
    """Substitute audited face truth into failing headline cells on the output sheets.
    Returns the list of overrides (for logging / the manifest). Saves in place."""
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.comments import Comment

    from app.engines.extraction.pipeline.template_map import _KEY_METRICS, _row_metric
    from app.engines.extraction.services.face_truth import build_face_truth
    from app.engines.extraction.services.formula_eval import evaluate
    from app.engines.extraction.services.formula_repair import _year_columns
    from app.engines.extraction.services.validation import _src_str

    face = build_face_truth(company.tables)
    if not face or not output_sheets:
        return []
    wb = load_workbook(workbook_path, data_only=False)
    overrides: list[Override] = []
    for title in output_sheets:
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        header_row, band = _year_columns(ws)
        if not band:
            continue
        year_of = {c: int(ws.cell(header_row, c).value) for c in band}
        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not isinstance(label, str) or not label.strip():
                continue
            cm = _row_metric(label.strip())
            if cm not in _KEY_METRICS:
                continue
            # Polarity. For UNAMBIGUOUS metrics the fixed convention is AUTHORITATIVE — an
            # asset / total / revenue is never negative on a face statement, an expense
            # always is — and must NOT be read from a sibling formula, because that formula
            # evaluates negative when the breakdown leaves are mis-extracted (which emitted
            # a negative 'Total Non-Current Assets'). Only sign-AMBIGUOUS metrics (profit
            # subtotals, other income — can be a loss / net expense) infer the sign from a
            # sibling formula, falling back to face truth's own sign. This keeps the P&L
            # additive AND preserves genuine losses, without flipping assets negative. #2/#6
            if cm in _ALWAYS_NEGATIVE:
                row_sign = -1.0
            elif cm in _ALWAYS_POSITIVE:
                row_sign = 1.0
            else:
                row_sign = None
                for c in band:
                    fc = ws.cell(r, c)
                    if isinstance(fc.value, str) and fc.value.startswith("="):
                        ev = evaluate(wb, title, fc.coordinate)
                        if ev is not None and abs(ev) > 1e-9:
                            row_sign = -1.0 if ev < 0 else 1.0
                            break
            for c in band:
                pair = face.get((cm, year_of[c]))
                if not pair:
                    continue
                truth, src = pair
                cell = ws.cell(r, c)
                if isinstance(cell, MergedCell):     # non-anchor of a merge — read-only
                    continue
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    computed = evaluate(wb, title, cell.coordinate)
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    computed = float(v)
                else:
                    computed = None
                # Write with the statement's intended sign (magnitude from face truth).
                value = row_sign * abs(truth) if row_sign is not None else float(truth)
                # A HEADLINE statement should show the audited figure exactly, not a
                # within-tolerance breakdown sum. Overriding every key cell to face truth
                # also makes the balance sheet tie (assets == equity+liabilities, since
                # both totals come from the same audited statement). Skip only when the
                # cell already holds that exact value (nothing to change).
                if computed is not None and abs(computed - value) <= 0.5:
                    continue
                was = cell.value
                cell.value = value
                if annotate:
                    prior = f"{computed:,.0f}" if computed is not None else "blank / not evaluable"
                    cell.comment = Comment(
                        f"OVERRIDE: substituted audited {cm} = {value:,.0f} (was {prior}). "
                        f"Source: {_src_str(src)}", "validation")
                pages = getattr(src, "pages", None) or []
                overrides.append(Override(title, cell.coordinate, cm, year_of[c], was,
                                          value, _src_str(src),
                                          report_file=getattr(src, "report_file", None),
                                          report_year=getattr(src, "report_year", None),
                                          page=(pages[0] if pages else None),
                                          table_id=getattr(src, "table_id", None)))
    if overrides:
        wb.save(workbook_path)
    wb.close()
    return overrides
