"""P4 — ValidatedOutputLedger.

Evaluates emitted key metrics (template plan OR no-template tables) against the
audited TrustedFaceIndex, produces an auditable ledger, and decides whether the
artifact is production-ready. A failing run is surfaced, not silently shipped.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Optional

from app.core.logging import get_logger
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.mapping import MappingPlan
from app.engines.extraction.services.face_truth import build_face_truth

logger = get_logger(__name__)

_LEDGER_SHEET = "Validation Ledger"
_HEADERS = ["Status", "Sheet", "Cell/Label", "Metric", "Year", "Value", "Face truth", "Source", "Note"]


@dataclass
class LedgerRow:
    status: str          # ok | fallback | sign | WITHHELD | MISMATCH
    sheet: str
    where: str           # cell coordinate or row label
    metric: str
    year: Optional[int]
    value: Optional[float]
    face: Optional[float]
    source: str
    note: str

    def to_row(self) -> list:
        return [self.status, self.sheet, self.where, self.metric or "", self.year,
                self.value, self.face, self.source, self.note]


def _src_str(src) -> str:
    if not src:
        return ""
    rep = getattr(src, "report_file", "") or ""
    pages = getattr(src, "pages", None) or []
    tid = getattr(src, "table_id", None)
    return f"{rep} p{pages}" + (f" [{tid}]" if tid else "")


def template_ledger(plan: MappingPlan) -> list[LedgerRow]:
    """Ledger rows for a template run: every withheld value + a sample of writes."""
    rows: list[LedgerRow] = []
    for cw in plan.withheld:
        rows.append(LedgerRow("WITHHELD", cw.sheet, cw.coordinate, cw.matched_label,
                              cw.year, cw.value, None, "", cw.note or ""))
    for cw in plan.writes:
        note = cw.note or ""
        status = "fallback" if note.startswith("fallback") else ("sign" if note.startswith("sign") else "ok")
        if status != "ok":   # surface the noteworthy writes (fallbacks, sign-corrections)
            rows.append(LedgerRow(status, cw.sheet, cw.coordinate, cw.matched_label,
                                  cw.year, cw.value, None, "", note))
    return rows


def no_template_ledger(company: CompanyResult, tieout) -> tuple[list[LedgerRow], int]:
    """C3 — validate the no-template output's key metrics against face truth.

    Returns (rows, mismatch_count). Every line resolving to a headline metric is
    checked against the audited face value for that year."""
    from app.engines.extraction.pipeline.template_map import _KEY_METRICS
    face = build_face_truth(company.tables)
    rows: list[LedgerRow] = []
    mismatches = 0
    for t in company.tables:
        for li in t.line_items:
            cm = li.canonical_metric
            if cm not in _KEY_METRICS:
                continue
            for v in li.values:
                if v.year is None or v.value is None:
                    continue
                truth_pair = face.get((cm, v.year))
                if not truth_pair:
                    continue
                truth, src = truth_pair
                ok = tieout(v.value, truth)
                if not ok:
                    mismatches += 1
                rows.append(LedgerRow(
                    "ok" if ok else "MISMATCH", t.title or t.statement_type.value,
                    li.label, cm, v.year, v.value, truth, _src_str(v.source or src),
                    "" if ok else "does not tie out to audited face statement",
                ))
    return rows, mismatches


def computed_output_ledger(workbook_path, company: CompanyResult, tieout,
                           output_sheets: set | None = None,
                           annotate: bool = True) -> tuple[list[LedgerRow], int, int]:
    """#1 — evaluate the WRITTEN workbook's formula rows (output P&L / balance sheet
    and breakdown subtotals) and reconcile any that resolve to a headline metric to
    the audited face truth. Catches frozen-reference formulas and wrong computed
    totals that the leaf-level plan tie-out never sees.

    Returns (rows, fail_count, unevaluable_count). A key-metric formula the evaluator
    can't parse is counted as UNEVALUATED (surfaced, never silently treated as a
    pass). When `annotate`, MISMATCH/UNEVALUATED cells get an in-cell comment so the
    workbook flags untrusted values without destroying the template's formula.

    Sign convention is scoped explicitly: a formula is validated SIGNED when it sits
    on an OUTPUT statement (`output_sheets` = the template's computed/formula sheets)
    OR pulls cross-sheet (`'PL1'!B29`); otherwise it's a breakdown subtotal that
    keeps its own positive-magnitude convention and is compared by MAGNITUDE."""
    from openpyxl import load_workbook
    from openpyxl.comments import Comment

    from app.engines.extraction.pipeline.template_map import _KEY_METRICS, _row_metric
    from app.engines.extraction.services.face_truth import build_face_truth
    from app.engines.extraction.services.formula_eval import evaluate

    face = build_face_truth(company.tables)
    if not face:
        return [], 0, 0
    output_sheets = output_sheets or set()
    wb = load_workbook(workbook_path, data_only=False)
    rows: list[LedgerRow] = []
    fails = 0
    unevaluable = 0
    dirty = False
    for ws in wb.worksheets:
        # year-header columns (first row in the top 8 with >=2 year-like ints)
        best: dict[int, int] = {}
        hdr = 1
        for r in range(1, min(ws.max_row, 8) + 1):
            cols = {c: int(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)
                    if isinstance(ws.cell(r, c).value, (int, float)) and 1990 <= int(ws.cell(r, c).value) <= 2100}
            if len(cols) > len(best):
                best, hdr = cols, r
        if len(best) < 2:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not isinstance(label, str) or not label.strip():
                continue
            cm = _row_metric(label.strip())
            if cm not in _KEY_METRICS:
                continue
            for c, year in best.items():
                cell = ws.cell(r, c)
                # only evaluate FORMULA cells (plain values were already plan-validated)
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                truth_pair = face.get((cm, year))
                if not truth_pair:
                    continue
                truth = truth_pair[0]
                computed = evaluate(wb, ws.title, cell.coordinate)
                if computed is None:
                    # Could not evaluate this formula -> surface it (NOT a silent pass).
                    unevaluable += 1
                    rows.append(LedgerRow("UNEVALUATED", ws.title, cell.coordinate, cm, year,
                                          None, truth, _src_str(truth_pair[1]),
                                          "formula not evaluable -> not validated"))
                    if annotate:
                        cell.comment = Comment(f"NOT VALIDATED: formula could not be evaluated "
                                               f"(audited {cm} = {truth:,.0f})", "validation")
                        dirty = True
                    continue
                # Signed on an output statement (whole-sheet) or any cross-sheet pull;
                # magnitude on intra-sheet breakdown subtotals (own positive convention).
                signed = ws.title in output_sheets or "!" in cell.value
                ok = tieout(computed, truth) if signed else tieout(abs(computed), abs(truth))
                if not ok:
                    fails += 1
                    if annotate:
                        cell.comment = Comment(f"VALIDATION FAILED: computed {computed:,.0f} does not "
                                               f"tie out to audited {cm} = {truth:,.0f}", "validation")
                        dirty = True
                rows.append(LedgerRow(
                    "ok" if ok else "MISMATCH", ws.title, cell.coordinate, cm, year,
                    round(computed, 2), truth, _src_str(truth_pair[1]),
                    "" if ok else "computed formula does not tie out to audited face statement",
                ))
    if dirty:
        wb.save(workbook_path)
    wb.close()
    return rows, fails, unevaluable


def headline_coverage_gaps(workbook_path, company: CompanyResult,
                           output_sheets: set) -> tuple[list[LedgerRow], int]:
    """Coverage gate — an emitted HEADLINE metric with NO audited face truth is an
    unvalidated value, NOT a silent pass.

    On the output sheets, every row resolving to a KEY metric is checked for face-truth
    coverage in each *historical* reporting year (forecast columns are exempt). A
    populated headline cell whose (metric, year) has no face truth is recorded as
    NO_FACE_TRUTH and counted as a production-blocking gap. This catches the case where
    a whole statement (e.g. a mis-classified balance sheet) produced no primary face
    table, so its output would otherwise pass review by never being validated at all."""
    from openpyxl import load_workbook

    from app.engines.extraction.pipeline.template_map import _KEY_METRICS, _row_metric
    from app.engines.extraction.services.formula_repair import _year_columns

    face = build_face_truth(company.tables)
    fyears = set(company.fiscal_years or [])
    if not output_sheets or not fyears:
        return [], 0
    wb = load_workbook(workbook_path, data_only=False)
    rows: list[LedgerRow] = []
    gaps = 0
    for title in output_sheets:
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        header_row, band = _year_columns(ws)
        if not band:
            continue
        year_of = {c: int(ws.cell(header_row, c).value) for c in band}
        seen: set = set()
        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not isinstance(label, str) or not label.strip():
                continue
            cm = _row_metric(label.strip())
            if cm not in _KEY_METRICS:
                continue
            for c in band:
                year = year_of[c]
                if year not in fyears:                 # only historical years, not forecasts
                    continue
                cell = ws.cell(r, c)
                has_content = (isinstance(cell.value, str) and cell.value.startswith("=")) \
                    or (isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool))
                if not has_content or face.get((cm, year)) is not None:
                    continue
                if (cm, year) in seen:
                    continue
                seen.add((cm, year))
                gaps += 1
                rows.append(LedgerRow("NO_FACE_TRUTH", title, cell.coordinate, cm, year,
                                      None, None, "",
                                      "emitted headline metric has no audited face truth -> unvalidated"))
    wb.close()
    return rows, gaps


_SOURCE_HEADERS = ["Sheet", "Cell", "Template label", "Matched label", "Year", "Value",
                   "Report year", "Report file", "Page", "Table id", "Confidence", "Note"]


def write_source_ledger(workbook_path, plan: MappingPlan, overrides=None) -> None:
    """Traceability (#9): a 'Source Ledger' sheet mapping every written cell back to
    its report / page / table of origin. Also records the FINAL output-cell overrides
    (audited face truth substituted into headline P&L / Balance Sheet cells), so the
    delivered output values are cell-level auditable, not only traceable via comments."""
    from openpyxl import load_workbook

    from app.engines.extraction.services import styles as S
    wb = load_workbook(workbook_path)
    name = "Source Ledger"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for c, h in enumerate(_SOURCE_HEADERS, start=1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill = S.HEADER_FONT, S.HEADER_FILL
    r = 2
    for w in sorted(plan.writes, key=lambda x: (x.sheet, x.coordinate)):
        src = w.source
        pages = (src.pages if src else None) or []
        vals = [w.sheet, w.coordinate, w.template_label, w.matched_label, w.year, w.value,
                w.source_report_year, getattr(src, "report_file", None),
                (pages[0] if pages else None), getattr(src, "table_id", None),
                round(w.confidence, 3), w.note]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v)
        r += 1
    # Final output overrides (headline cells substituted with audited face truth).
    for o in sorted(overrides or [], key=lambda x: (x.sheet, x.coordinate)):
        vals = [o.sheet, o.coordinate, o.metric, "(headline override)", o.year, o.value,
                getattr(o, "report_year", None), getattr(o, "report_file", None) or o.source,
                getattr(o, "page", None), getattr(o, "table_id", None), 1.0,
                f"OVERRIDE: audited face truth (was {o.was!r})"]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c, v)
        r += 1
    ws.freeze_panes = "A2"
    wb.save(workbook_path)


def write_scope_note(workbook_path, *, cash_flow_in_scope: bool,
                     detail_incomplete_sheets: list) -> None:
    """Add a first-tab 'Scope & Notes' sheet that explicitly states the delivery scope —
    headline P&L/BS are audited-exact; cash flow is in/out of scope; and which detail
    schedules carry MATERIAL unmapped detail. Makes the workbook self-describing so a
    reader never mistakes a supporting note for a fully-mapped, reliable schedule."""
    from openpyxl import load_workbook

    from app.engines.extraction.services import styles as S
    wb = load_workbook(workbook_path)
    name = "Scope & Notes"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)        # first tab -> most visible
    cf = ("IN SCOPE" if cash_flow_in_scope
          else "OUT OF SCOPE — the template defines no cash-flow output sheet.")
    lines = [
        ("DELIVERY SCOPE", True),
        ("Headline statements (Profit & Loss, Balance Sheet) tie EXACTLY to the audited "
         "PDF face statements for all reporting years.", False),
        ("", False),
        (f"Cash Flow: {cf}", False),
        ("", False),
        ("Detail schedules (BS1–BS5, PL1–PL7) are SUPPORTING notes — not all are fully "
         "mapped from the source. Subtotals off by >5% are NOT plugged; they are disclosed "
         "as DETAIL_INCOMPLETE in the 'Validation Ledger'. Use headline statements for "
         "reliance; treat detail tabs as indicative where flagged.", False),
        ("", False),
        ("Sheets with MATERIAL unmapped detail:" if detail_incomplete_sheets
         else "No sheet carries material unmapped detail.", True),
    ]
    r = 1
    for text, bold in lines:
        cell = ws.cell(r, 1, text)
        if bold:
            cell.font = S.HEADER_FONT
        r += 1
    for sh in detail_incomplete_sheets:
        ws.cell(r, 1, f"   • {sh}")
        r += 1
    ws.column_dimensions["A"].width = 110
    wb.save(workbook_path)


def _wb_fingerprint(path) -> tuple[int, int]:
    """(formula-cell count, styled-cell count) — used to verify a LibreOffice round-trip
    didn't drop content before we adopt the converted file."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    try:
        formulas = styled = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.data_type == "f" or (isinstance(c.value, str) and c.value.startswith("=")):
                        formulas += 1
                    if c.has_style:
                        styled += 1
        return formulas, styled
    finally:
        wb.close()


def recalc_workbook(workbook_path) -> bool:
    """Make formula cells readable by NON-Excel consumers. Always sets `fullCalcOnLoad`
    so Excel/LibreOffice recalculate on open; if a LibreOffice binary is found, headless-
    recalculates and rewrites the file so cached <v> values are materialized for openpyxl
    /pandas/`data_only` readers too. Returns True iff cached values were materialized."""
    import shutil
    import subprocess
    import tempfile
    from pathlib import Path

    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=False)   # keep formulas, never their cached values
    try:
        wb.calculation.fullCalcOnLoad = True   # Excel/LibreOffice recalc on open
    except Exception as exc:  # noqa: BLE001
        logger.debug("Could not set fullCalcOnLoad: %s", exc)
    wb.save(workbook_path)
    wb.close()

    soffice = (shutil.which("soffice") or shutil.which("libreoffice")
               or next((p for p in (
                   r"C:\Program Files\LibreOffice\program\soffice.exe",
                   r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
                   "/usr/bin/soffice", "/usr/bin/libreoffice",
                   "/Applications/LibreOffice.app/Contents/MacOS/soffice",
               ) if Path(p).exists()), None))
    if not soffice:
        logger.info("Formula cache: set fullCalcOnLoad (Excel recalcs on open); no LibreOffice "
                    "found, so cached values are NOT materialized for headless readers.")
        return False
    try:
        src = Path(workbook_path)
        with tempfile.TemporaryDirectory() as td:
            subprocess.run([soffice, "--headless", "--calc",
                            "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                            "--outdir", td, str(src)], check=True, timeout=180,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            out = Path(td) / (src.stem + ".xlsx")
            if out.exists():
                # LibreOffice's xlsx->xlsx round-trip can silently drop formulas/styles.
                # Only adopt the recalced file if it preserved EVERY formula and kept the
                # bulk of styled cells; otherwise keep the openpyxl file (fullCalcOnLoad
                # already set) — we never trade correctness for cached values.
                before_f, before_s = _wb_fingerprint(src)
                after_f, after_s = _wb_fingerprint(out)
                if after_f < before_f or (before_s and after_s < 0.9 * before_s):
                    logger.warning(
                        "LibreOffice recalc lost content (formulas %d->%d, styled cells "
                        "%d->%d); discarding it and keeping the openpyxl workbook with "
                        "fullCalcOnLoad set.", before_f, after_f, before_s, after_s)
                    return False
                shutil.copyfile(out, src)
                logger.info("Formula cache materialized via LibreOffice headless recalc.")
                return True
    except Exception as exc:  # noqa: BLE001 — recalc is best-effort, never fail the run
        logger.warning("LibreOffice recalc failed (%s); fullCalcOnLoad still set.", exc)
    return False


# Plug materiality (gap as a fraction of the audited total). A subtotal off by more than
# MATERIAL is NOT plugged — forcing a near-empty section to its audited total is fake
# precision; it's disclosed as DETAIL_INCOMPLETE instead. Small gaps (rounding / a little
# unmapped leaf detail) are plugged so the schedule still foots, flagged minor/warning.
_PLUG_MINOR = 0.01     # <=1% gap: rounding -> plug, minor
_PLUG_MATERIAL = 0.05  # >5% gap: material -> DO NOT plug, disclose as DETAIL_INCOMPLETE


def reconcile_breakdown_subtotals(workbook_path, company: CompanyResult, tieout,
                                  output_sheets: set, annotate: bool = True) -> tuple[list[LedgerRow], int]:
    """Reconcile each BREAKDOWN-sheet subtotal to its audited total, MATERIALITY-GATED and
    honestly graded — never forcing a near-empty section to match (no fake precision).

    The audited reference is the DetailTruthIndex (the schedules' OWN note totals; primary
    face totals fall back only when a note total is absent), so detail is checked against
    detail, not against a primary total masquerading as detail.

    Per subtotal that doesn't already tie (those are MAPPED_OK in the computed ledger):
      * gap <= 5% of audited -> append a literal plug so the schedule foots; status
        DETAIL_PLUG (minor if <=1%, else warning), variance disclosed in the comment;
      * gap > 5% -> NOT plugged; status DETAIL_INCOMPLETE (material) with the variance
        disclosed — the leaves are genuinely unmapped and we say so.
    Output sheets are skipped (headline override handles them). Returns (rows, plug_count);
    DETAIL_INCOMPLETE rows are in `rows` for the manifest to count separately."""
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.comments import Comment

    from app.engines.extraction.pipeline.template_map import _row_metric
    from app.engines.extraction.services.formula_eval import evaluate
    from app.engines.extraction.services.formula_repair import _year_columns

    face = build_face_truth(company.tables)
    detail_truth = build_face_truth(company.tables, notes_only=True)   # DetailTruthIndex
    if not face and not detail_truth:
        return [], 0
    wb = load_workbook(workbook_path, data_only=False)
    rows: list[LedgerRow] = []
    reconciled = 0
    dirty = False
    for ws in wb.worksheets:
        if ws.title in output_sheets:        # breakdown/note sheets only
            continue
        header_row, band = _year_columns(ws)
        if not band:
            continue
        year_of = {c: int(ws.cell(header_row, c).value) for c in band}
        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not isinstance(label, str) or not label.strip():
                continue
            cm = _row_metric(label.strip())
            if cm is None:
                continue
            for c in band:
                cell = ws.cell(r, c)
                if isinstance(cell, MergedCell):
                    continue
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue                  # only aggregate FORMULA subtotals
                # DetailTruthIndex first (note's own total); fall back to face for grand
                # totals that only exist on the primary statement.
                pair = detail_truth.get((cm, year_of[c])) or face.get((cm, year_of[c]))
                if not pair:
                    continue
                audited = abs(pair[0])        # breakdown positive-magnitude convention
                ev = evaluate(wb, ws.title, cell.coordinate)
                if ev is None or tieout(abs(ev), audited):
                    continue                  # unevaluable, or already ties (MAPPED_OK)
                gap = audited - abs(ev)
                pct = abs(gap) / audited if audited else 1.0
                if pct > _PLUG_MATERIAL:
                    # Material -> do NOT plug; disclose the unmapped detail honestly.
                    if annotate:
                        cell.comment = Comment(
                            f"DETAIL INCOMPLETE: mapped leaves = {abs(ev):,.0f} vs audited "
                            f"{cm} = {audited:,.0f} ({pct:.0%} unmapped) — NOT plugged.", "validation")
                    rows.append(LedgerRow("DETAIL_INCOMPLETE", ws.title, cell.coordinate, cm,
                                          year_of[c], round(audited, 2), round(abs(ev), 2), "",
                                          f"{pct:.0%} unmapped (MATERIAL) — left unreconciled"))
                    continue
                plug = audited - ev           # orig + plug evaluates to the audited magnitude
                sev = "minor" if pct <= _PLUG_MINOR else "warning"
                cell.value = f"{cell.value}+{plug:.2f}" if plug >= 0 else f"{cell.value}-{abs(plug):.2f}"
                if annotate:
                    cell.comment = Comment(
                        f"DETAIL PLUG ({sev}, {pct:.1%}): mapped leaves = {abs(ev):,.0f}; "
                        f"+{gap:,.0f} unmapped to reach audited {cm} = {audited:,.0f}", "validation")
                rows.append(LedgerRow("DETAIL_PLUG", ws.title, cell.coordinate, cm, year_of[c],
                                      round(audited, 2), round(abs(ev), 2), "",
                                      f"plugged {gap:,.0f} ({pct:.1%}, {sev})"))
                reconciled += 1
                dirty = True
    if dirty:
        wb.save(workbook_path)
    wb.close()
    return rows, reconciled


def identity_ledger(company: CompanyResult) -> tuple[list[LedgerRow], int]:
    """Accounting-identity consistency of the audited face truth (P&L waterfall + BS
    composition). Catches face-truth EXTRACTION errors that the external tie-out can't
    see (we override output to face, so output==face passes while face itself is wrong).
    Advisory: surfaced as IDENTITY_OK/IDENTITY_FAIL rows; failures count toward the
    strict `fully_reconciled` flag, not the headline production gate."""
    from app.engines.extraction.services.identity_checks import check_identities, check_sign_sanity

    face = build_face_truth(company.tables)
    findings = check_identities(face, company.fiscal_years) \
        + check_sign_sanity(face, company.fiscal_years)
    rows: list[LedgerRow] = []
    fails = 0
    for f in findings:
        if not f.ok:
            fails += 1
        if f.ok:
            note = ""
        elif f.statement == "SANITY":
            note = "emitted value is negative on a must-be-positive metric (sign/extraction error)"
        else:
            note = "face-truth values do not satisfy this accounting identity"
        rows.append(LedgerRow(
            "IDENTITY_OK" if f.ok else "IDENTITY_FAIL", f.statement, f.name, "", f.year,
            round(f.actual, 2), round(f.expected, 2), "", note,
        ))
    return rows, fails


def widen_columns(workbook_path) -> None:
    """Fit each column to its widest value so large figures don't render as '######'.
    Year data columns ship from the template at the default ~8.4 chars, too narrow for
    9-digit money. Bounded [10, 60] and never shrinks an existing (wider) width."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path, data_only=False)   # explicit: preserve formulas on save
    for ws in wb.worksheets:
        widest: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                s = f"{int(round(v)):,}" if isinstance(v, (int, float)) and not isinstance(v, bool) else str(v)
                col = get_column_letter(cell.column)
                if len(s) > widest.get(col, 0):
                    widest[col] = len(s)
        for col, w in widest.items():
            cur = ws.column_dimensions[col].width or 0
            ws.column_dimensions[col].width = min(max(cur, w + 2, 10), 60)
    wb.save(workbook_path)


def append_ledger_sheet(workbook_path, rows: list[LedgerRow]) -> None:
    """Append a 'Validation Ledger' sheet to an existing workbook."""
    from openpyxl import load_workbook

    from app.engines.extraction.services import styles as S
    wb = load_workbook(workbook_path)
    if _LEDGER_SHEET in wb.sheetnames:
        del wb[_LEDGER_SHEET]
    ws = wb.create_sheet(_LEDGER_SHEET)
    for c, name in enumerate(_HEADERS, start=1):
        cell = ws.cell(1, c, name)
        cell.font, cell.fill = S.HEADER_FONT, S.HEADER_FILL
    for r, row in enumerate(sorted(rows, key=lambda x: (x.status == "ok", x.sheet)), start=2):
        for c, val in enumerate(row.to_row(), start=1):
            ws.cell(r, c, val)
    ws.freeze_panes = "A2"
    wb.save(workbook_path)
