"""Dependency-free evaluator for the formula grammar these templates use.

Handles: numbers, cell refs (with optional `$` and `'Sheet'!`/`Sheet!` prefix),
ranges inside SUM(...), unary/binary `+`/`-`, and parentheses. Anything outside
this grammar (other functions, `*`/`/`, etc.) raises `_Bail`, and the caller
treats the cell as un-evaluable (skipped — never a false validation failure).

Used by the template computed-formula tie-out (#1): it lets us evaluate output /
subtotal FORMULA rows (PAT, total assets, revenue) from the leaf values we wrote
and reconcile them to the audited face statements — catching frozen-reference
formulas and wrong computed totals that plain leaf checks miss.
"""
from __future__ import annotations

import re

from openpyxl.utils import column_index_from_string, get_column_letter

_TOKEN = re.compile(
    r"\s*(?:"
    r"(?P<num>\d+(?:\.\d+)?)"
    r"|(?P<ref>(?:(?:'[^']+'|[A-Za-z0-9_.\-& ]+)!)?\$?[A-Za-z]+\$?\d+)"
    r"|(?P<op>[()+\-:,])"
    r"|(?P<fn>[A-Za-z]+)\("
    r")"
)


class _Bail(Exception):
    """Formula uses a construct we don't evaluate — skip this cell."""


def _tokenize(expr: str):
    toks, i = [], 0
    while i < len(expr):
        if expr[i].isspace():
            i += 1
            continue
        m = _TOKEN.match(expr, i)
        if not m or m.end() == i:
            raise _Bail(f"unparseable at {expr[i:]!r}")
        if m.group("fn"):
            toks.append(("fn", m.group("fn").upper()))
        elif m.group("num"):
            toks.append(("num", float(m.group("num"))))
        elif m.group("ref"):
            toks.append(("ref", m.group("ref")))
        else:
            toks.append(("op", m.group("op")))
        i = m.end()
    return toks


def _split_ref(ref: str) -> tuple[str | None, str]:
    """'PL1 - Revenue'!$B$5 -> ('PL1 - Revenue', 'B5')."""
    sheet = None
    if "!" in ref:
        sheet, ref = ref.rsplit("!", 1)
        sheet = sheet.strip().strip("'")
    return sheet, ref.replace("$", "")


def _coord_parts(coord: str) -> tuple[int, int]:
    m = re.match(r"([A-Za-z]+)(\d+)", coord)
    return column_index_from_string(m.group(1).upper()), int(m.group(2))


class _Evaluator:
    def __init__(self, wb, default_sheet, cache):
        self.wb, self.default_sheet, self.cache = wb, default_sheet, cache

    def cell_value(self, sheet: str | None, coord: str) -> float:
        title = sheet or self.default_sheet
        if (title, coord) in self.cache:
            v = self.cache[(title, coord)]
            if v is _IN_PROGRESS:
                raise _Bail("circular reference")
            return v
        self.cache[(title, coord)] = _IN_PROGRESS
        try:
            ws = self.wb[title]
        except KeyError:
            raise _Bail(f"unknown sheet {title!r}")
        raw = ws[coord].value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            out = 0.0
        elif isinstance(raw, (int, float)):
            out = float(raw)
        elif isinstance(raw, str) and raw.startswith("="):
            out = _Evaluator(self.wb, title, self.cache)._eval(_tokenize(raw[1:]))
        else:
            out = 0.0  # text cell
        self.cache[(title, coord)] = out
        return out

    def _range_sum(self, a: str, b: str, sheet: str | None) -> float:
        c1, r1 = _coord_parts(a)
        c2, r2 = _coord_parts(b)
        total = 0.0
        for c in range(min(c1, c2), max(c1, c2) + 1):
            for r in range(min(r1, r2), max(r1, r2) + 1):
                total += self.cell_value(sheet, f"{get_column_letter(c)}{r}")
        return total

    def _eval(self, toks: list) -> float:
        self.toks, self.pos = toks, 0
        val = self._expr()
        if self.pos != len(self.toks):
            raise _Bail("trailing tokens")
        return val

    def _peek(self):
        return self.toks[self.pos] if self.pos < len(self.toks) else (None, None)

    def _expr(self) -> float:
        val = self._term()
        while self._peek() == ("op", "+") or self._peek() == ("op", "-"):
            op = self.toks[self.pos][1]; self.pos += 1
            rhs = self._term()
            val = val + rhs if op == "+" else val - rhs
        return val

    def _term(self) -> float:
        kind, v = self._peek()
        if (kind, v) == ("op", "-"):
            self.pos += 1
            return -self._term()
        if (kind, v) == ("op", "+"):
            self.pos += 1
            return self._term()
        return self._factor()

    def _factor(self) -> float:
        kind, v = self._peek()
        if kind == "num":
            self.pos += 1
            return v
        if kind == "fn":
            if v != "SUM":
                raise _Bail(f"unsupported function {v}")
            self.pos += 1
            return self._sum_args()
        if (kind, v) == ("op", "("):
            self.pos += 1
            val = self._expr()
            if self._peek() != ("op", ")"):
                raise _Bail("missing )")
            self.pos += 1
            return val
        if kind == "ref":
            self.pos += 1
            sheet, coord = _split_ref(v)
            # range outside SUM is not supported
            if self._peek() == ("op", ":"):
                raise _Bail("bare range")
            return self.cell_value(sheet, coord)
        raise _Bail(f"unexpected token {kind} {v}")

    def _sum_args(self) -> float:
        total = 0.0
        while True:
            kind, v = self._peek()
            if kind == "ref":
                self.pos += 1
                sheet, coord = _split_ref(v)
                if self._peek() == ("op", ":"):
                    self.pos += 1
                    k2, v2 = self._peek()
                    if k2 != "ref":
                        raise _Bail("bad range")
                    self.pos += 1
                    _s2, coord2 = _split_ref(v2)
                    total += self._range_sum(coord, coord2, sheet)
                else:
                    total += self.cell_value(sheet, coord)
            else:
                total += self._expr()
            kind, v = self._peek()
            if (kind, v) == ("op", ","):
                self.pos += 1
                continue
            if (kind, v) == ("op", ")"):
                self.pos += 1
                return total
            raise _Bail("bad SUM args")


_IN_PROGRESS = object()


def evaluate(wb, sheet_title: str, coord: str) -> float | None:
    """Evaluate a workbook cell (following formulas) or None if un-evaluable."""
    try:
        return _Evaluator(wb, sheet_title, {}).cell_value(sheet_title, coord)
    except _Bail:
        return None
    except Exception:  # noqa: BLE001 — never let a parse edge case crash the run
        return None
