"""Tests for Layer 5 template mapping."""
from pathlib import Path

import openpyxl
import pytest

from app.core.config import get_settings
from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.pipeline.template_map import (
    _best_candidate, _match_norm, _related_types, apply_plan, build_plan,
)


@pytest.fixture(autouse=True)
def _no_embeddings(monkeypatch):
    get_settings.cache_clear()
    monkeypatch.setenv("USE_EMBEDDINGS", "false")
    yield
    get_settings.cache_clear()


def _revenue_company() -> CompanyResult:
    table = FinancialTable(
        statement_type=StatementType.revenue, title="Revenue",
        line_items=[
            LineItem(label="Local sales", values=[
                LineItemValue(year=2024, value=50.0, source_report_year=2025),
                LineItemValue(year=2025, value=60.0, source_report_year=2025)]),
            LineItem(label="Export sales", values=[
                LineItemValue(year=2024, value=30.0),
                LineItemValue(year=2025, value=40.0)]),
        ],
    )
    return CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[table])


def _make_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    pl = wb.active
    pl.title = "PL1 - Revenue"
    pl["A1"] = "Note 27 - Gross Revenue"
    pl["B2"], pl["D2"] = "Historic", "Forecasted"
    pl["A3"], pl["B3"], pl["C3"], pl["D3"] = "Particulars", 2024, 2025, 2026  # D = forecast
    pl["A4"] = "GROSS REVENUE"                       # section header
    pl["A5"] = "Local sales"                         # leaf (empty)
    pl["A6"] = "Export sales"                        # leaf (empty)
    pl["A7"], pl["B7"], pl["C7"] = "Gross Revenue", "=SUM(B5:B6)", "=SUM(C5:C6)"  # subtotal

    out = wb.create_sheet("P&L")                     # output sheet: all formulas
    out["A1"] = "Income Statement"
    out["B3"], out["C3"], out["D3"] = 2024, 2025, 2026
    out["D2"] = "Forecasted"
    out["A5"], out["B5"], out["C5"] = "Revenue", "='PL1 - Revenue'!B5", "='PL1 - Revenue'!C5"

    mgmt = wb.create_sheet("Mgmt Info.")             # no year header
    mgmt["A1"] = "Company profile"
    wb.save(path)


def test_build_plan_targets_leaf_historical_cells(tmp_path):
    tpl = tmp_path / "template.xlsx"
    _make_template(tpl)
    plan = build_plan(_revenue_company(), tpl)

    assert "PL1 - Revenue" in plan.sheets_processed
    assert "P&L" in plan.sheets_skipped          # formulas -> low empty fraction
    assert "Mgmt Info." in plan.sheets_skipped   # no year header

    cells = {(w.coordinate, w.value) for w in plan.writes}
    assert ("B5", 50.0) in cells and ("C5", 60.0) in cells   # Local sales 2024/2025
    assert ("B6", 30.0) in cells and ("C6", 40.0) in cells   # Export sales
    # Never write forecast (col D) or formula cells (B7/C7).
    coords = {w.coordinate for w in plan.writes}
    assert not any(c.startswith("D") for c in coords)
    assert "B7" not in coords and "C7" not in coords


def test_apply_plan_preserves_formulas_and_writes_values(tmp_path):
    tpl = tmp_path / "template.xlsx"
    _make_template(tpl)
    out = tmp_path / "out.xlsx"
    apply_plan(build_plan(_revenue_company(), tpl), tpl, out)

    wb = openpyxl.load_workbook(out, data_only=False)
    pl = wb["PL1 - Revenue"]
    assert pl["B5"].value == 50.0 and pl["C5"].value == 60.0
    assert pl["B7"].value == "=SUM(B5:B6)"     # subtotal formula preserved
    assert pl["D5"].value is None              # forecast column untouched
    assert wb["P&L"]["B5"].value == "='PL1 - Revenue'!B5"  # output sheet untouched
    wb.close()


# --- lever 1: statement-family containment ---

def test_related_types_widens_within_family_only():
    rel = _related_types(StatementType.current_assets)
    assert StatementType.balance_sheet in rel          # parent face
    assert StatementType.current_liabilities in rel    # sibling breakdown
    assert StatementType.revenue not in rel            # different family -> no bleed
    assert _related_types(None) == set()


def _split_company() -> CompanyResult:
    # "Local sales" lives in the revenue (own) table; "Export sales" only in the
    # income_statement (parent face) table -> reachable only via family widening.
    rev = FinancialTable(
        statement_type=StatementType.revenue, title="Revenue",
        line_items=[LineItem(label="Local sales", values=[
            LineItemValue(year=2024, value=50.0), LineItemValue(year=2025, value=60.0)])],
    )
    face = FinancialTable(
        statement_type=StatementType.income_statement, title="Income Statement",
        line_items=[LineItem(label="Export sales", values=[
            LineItemValue(year=2024, value=30.0), LineItemValue(year=2025, value=40.0)])],
    )
    return CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[rev, face])


def test_family_widening_maps_line_from_parent_face_table(tmp_path):
    tpl = tmp_path / "template.xlsx"
    _make_template(tpl)  # PL1 - Revenue sheet -> classifies to `revenue`
    plan = build_plan(_split_company(), tpl)
    cells = {(w.coordinate, w.value) for w in plan.writes}
    assert ("B5", 50.0) in cells                       # own-type (revenue) match
    assert ("B6", 30.0) in cells and ("C6", 40.0) in cells  # widened to income_statement face


# --- lever 2: abbreviation-aware matching ---

def test_match_norm_expands_corporate_abbreviations():
    assert _match_norm("Hyundai Nishat Motors (Pvt) Ltd") == \
        _match_norm("Hyundai Nishat Motors (Private) Limited")


def test_abbrev_lifts_fuzzy_match_over_threshold():
    from rapidfuzz import fuzz
    a = _match_norm("Dividend from XYZ Co Ltd")
    b = _match_norm("Dividend from XYZ Company Limited")
    assert fuzz.token_set_ratio(a, b) >= 82


def _two_sheet_template(path: Path) -> None:
    # Two writable breakdown sheets that both contain the same row label.
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "PL1 - Revenue"
    a["A1"] = "Note - Revenue"
    a["A3"], a["B3"], a["C3"] = "Particulars", 2024, 2025
    a["A4"] = "Shared item"
    a["A5"] = "Filler one"
    b = wb.create_sheet("PL4 - Other Income")
    b["A1"] = "Note - Other Income"
    b["A3"], b["B3"], b["C3"] = "Particulars", 2024, 2025
    b["A4"] = "Shared item"
    b["A5"] = "Filler two"
    wb.save(path)


def test_global_dedup_writes_shared_line_to_one_sheet_only(tmp_path):
    # One extracted line ('Shared item', in the revenue table) is reachable from both
    # sheets (own on PL1, widened on PL4). The global dedup must write it to exactly
    # ONE sheet, not both — preventing the cross-statement bleed.
    tpl = tmp_path / "two.xlsx"
    _two_sheet_template(tpl)
    company = CompanyResult(
        company="Acme", fiscal_years=[2024, 2025],
        tables=[FinancialTable(
            statement_type=StatementType.revenue, title="Revenue",
            line_items=[LineItem(label="Shared item", values=[
                LineItemValue(year=2024, value=11.0), LineItemValue(year=2025, value=22.0)])],
        )],
    )
    plan = build_plan(company, tpl)
    sheets_with_shared = {w.sheet for w in plan.writes if w.matched_label == "Shared item"}
    assert len(sheets_with_shared) == 1            # written to exactly one sheet, not both
    assert "Shared item" in plan.unmatched_template_labels  # the losing sheet's row recorded


# --- validation gate: semantic guards + tie-out ---

def test_metric_agreement_guard_rejects_wrong_concept():
    # A 'cash in hand' line must not fill a row whose metric is share capital.
    line = LineItem(label="Cash in hand", canonical_metric="cash_in_hand",
                    values=[LineItemValue(year=2025, value=2343.0)])
    cands = [(_match_norm("Cash in hand"), "", line, StatementType.current_assets)]
    best, _ = _best_candidate("Cash in hand", "", cands, 50.0, row_metric="share_capital")
    assert best is None                              # confident-but-different metric -> rejected
    best2, _ = _best_candidate("Cash in hand", "", cands, 50.0, row_metric=None)
    assert best2 is line                             # no row metric -> allowed (recall preserved)


def test_role_guard_rejects_total_into_leaf_but_allows_total_into_total():
    line = LineItem(label="Total revenue", role="total",
                    values=[LineItemValue(year=2025, value=999.0)])
    cands = [(_match_norm("Total revenue"), "", line, StatementType.income_statement)]
    # leaf template row -> total candidate rejected
    assert _best_candidate("Total revenue", "", cands, 50.0, template_is_total=False)[0] is None
    # total template row -> total candidate allowed (#7 fix)
    assert _best_candidate("Total revenue", "", cands, 50.0, template_is_total=True)[0] is line


def test_polarity_guard_rejects_cross_side_match():
    # An asset line (current_assets) must not fill an equity-polarity sheet row.
    line = LineItem(label="Cash in hand", values=[LineItemValue(year=2025, value=2343.0)])
    cands = [(_match_norm("Cash in hand"), "", line, StatementType.current_assets)]
    best, _ = _best_candidate("Cash in hand", "", cands, 50.0, sheet_polarity="equity")
    assert best is None                              # asset vs equity -> rejected (no metric needed)
    best2, _ = _best_candidate("Cash in hand", "", cands, 50.0, sheet_polarity="asset")
    assert best2 is line                             # same side -> allowed


def _tieout_template(path: Path) -> None:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PL1 - Revenue"
    ws["A1"] = "Revenue note"
    ws["A3"], ws["B3"], ws["C3"] = "Particulars", 2024, 2025   # needs >=2 year columns
    ws["A4"] = "Revenue from contracts"     # leaf to fill
    ws["A5"] = "Filler row"
    wb.save(path)


def test_tieout_gate_withholds_value_contradicting_face_statement(tmp_path):
    tpl = tmp_path / "t.xlsx"; _tieout_template(tpl)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        # Audited truth (label deliberately distinct so it doesn't compete for the row).
        FinancialTable(statement_type=StatementType.income_statement, title="P&L",
            line_items=[LineItem(label="Turnover net", canonical_metric="revenue",
                values=[LineItemValue(year=2025, value=1000.0)])]),
        # Breakdown line the template row matches — its 2025 value contradicts truth.
        FinancialTable(statement_type=StatementType.revenue, title="Revenue",
            line_items=[LineItem(label="Revenue from contracts", canonical_metric="revenue",
                values=[LineItemValue(year=2025, value=9999.0)])]),
    ])
    plan = build_plan(company, tpl)
    assert all(w.value != 9999.0 for w in plan.writes)                    # not shipped
    assert any(w.value == 9999.0 and "tieout" in (w.note or "") for w in plan.withheld)


def test_no_template_ledger_flags_face_mismatch():
    # P4/C3: a no-template key-metric line that contradicts the audited face truth
    # is flagged MISMATCH and counted (run becomes non-production).
    from app.engines.extraction.pipeline.template_map import _tieout
    from app.engines.extraction.services.validation import no_template_ledger
    company = CompanyResult(company="A", fiscal_years=[2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",  # -> primary face
                       line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                           values=[LineItemValue(year=2025, value=1000.0)])]),
        FinancialTable(statement_type=StatementType.revenue, title="Revenue note",  # -> note
                       line_items=[LineItem(label="Revenue from contracts", canonical_metric="revenue",
                           values=[LineItemValue(year=2025, value=9999.0)])]),
    ])
    rows, fails = no_template_ledger(company, _tieout)
    assert fails >= 1
    assert any(r.status == "MISMATCH" and r.value == 9999.0 for r in rows)
    assert any(r.status == "ok" and r.value == 1000.0 for r in rows)


_LUCKY = Path(r"C:\AI Financial Intelligence\data\lucky-template.xlsx")


@pytest.mark.skipif(not _LUCKY.exists(), reason="real Lucky template not available")
def test_real_lucky_template_maps_revenue_leaves(tmp_path):
    plan = build_plan(_revenue_company(), _LUCKY)
    # Output sheets are never written.
    assert "P&L" in plan.sheets_skipped and "Balance Sheet" in plan.sheets_skipped
    # Revenue leaves land on the PL1 sheet.
    pl1 = [w for w in plan.writes if w.sheet == "PL1 - Revenue"]
    assert any(w.matched_label == "Local sales" for w in pl1)
    assert any(w.matched_label == "Export sales" for w in pl1)
