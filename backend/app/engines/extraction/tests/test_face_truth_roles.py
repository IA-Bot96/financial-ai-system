"""Tests for table-role inference (P1/P3) — especially the analytical-vs-primary
precedence that decides whether a statement contributes face truth."""
from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.services.face_truth import build_face_truth, infer_table_role


def _bs(title, unit="thousands", with_totals=True):
    items = []
    if with_totals:
        items = [
            LineItem(label="Total assets", canonical_metric="total_assets",
                     canonical_category="balance_sheet",
                     values=[LineItemValue(year=2025, value=266748030.0)]),
            LineItem(label="Total equity and liabilities",
                     canonical_metric="total_equity_and_liabilities",
                     canonical_category="balance_sheet",
                     values=[LineItemValue(year=2025, value=266748030.0)]),
        ]
    return FinancialTable(statement_type=StatementType.balance_sheet, title=title,
                          unit_scale=unit, line_items=items)


def test_analysis_of_sofp_with_face_totals_is_primary():
    # The Lucky case: title contains "analysis" but it IS the statement (absolute
    # figures + headline totals) -> primary, so it can supply balance-sheet face truth.
    t = _bs("Analysis of Statement of Financial Position")
    assert infer_table_role(t) == "primary"


def test_real_sofp_is_primary():
    assert infer_table_role(_bs("Unconsolidated Statement of Financial Position")) == "primary"


def test_horizontal_and_vertical_analysis_stay_analytical():
    # Strong analytical markers win even with face totals present.
    assert infer_table_role(_bs("Horizontal Analysis of the Balance Sheet")) == "analytical"
    assert infer_table_role(_bs("Vertical Analysis of Financial Position")) == "analytical"


def test_six_year_summary_with_totals_stays_analytical():
    assert infer_table_role(_bs("Six Year Summary - Statement of Financial Position")) == "analytical"


def test_percent_scale_is_analytical_even_if_titled_like_a_statement():
    assert infer_table_role(_bs("Statement of Financial Position", unit="%")) == "analytical"


def test_bare_analysis_without_face_evidence_is_analytical():
    # A generic "analysis" block with no face title/totals is analytical, not a note.
    t = FinancialTable(statement_type=StatementType.non_current_assets,
                       title="Analysis of operations", unit_scale="thousands",
                       line_items=[LineItem(label="Something", canonical_metric="property_plant_equipment",
                                            canonical_category="balance_sheet",
                                            values=[LineItemValue(year=2025, value=1.0)])])
    assert infer_table_role(t) == "analytical"


def test_common_size_percentages_do_not_pollute_truth():
    # A mixed "Analysis of SoFP" tags BOTH the real figure AND its common-size % (100)
    # as total_assets, and the % rows can outnumber the real ones. The currency-scale
    # anchor must keep the real value, not collapse to 100.
    items = []
    real = {2022: 184962368.0, 2023: 213079067.0, 2024: 234018090.0, 2025: 266748030.0}
    for y, v in real.items():
        items.append(LineItem(label="Total assets", canonical_metric="total_assets",
                              canonical_category="balance_sheet", values=[LineItemValue(year=y, value=v)]))
        # two percentage rows per year (common-size + a ratio) — outnumber the real one
        for pct in (100.0, 13.99):
            items.append(LineItem(label="Total assets %", canonical_metric="total_assets",
                                  canonical_category="balance_sheet", values=[LineItemValue(year=y, value=pct)]))
    t = FinancialTable(statement_type=StatementType.balance_sheet,
                       title="Analysis of Statement of Financial Position", unit_scale="thousands",
                       line_items=items)
    truth = build_face_truth([t])
    assert truth[("total_assets", 2025)][0] == 266748030.0   # real value, not 100


def test_expense_metrics_normalised_negative():
    # Cost of sales reported POSITIVE in a note -> face truth stores it negative so it
    # matches the additive-P&L formula sign (=-'PL2'!..) and the signed tie-out passes.
    t = FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label="Cost of sales", canonical_metric="cost_of_sales",
                           canonical_category="income_statement",
                           values=[LineItemValue(year=2025, value=81827060.0)])])
    truth = build_face_truth([t])
    assert truth[("cost_of_sales", 2025)][0] == -81827060.0


def test_crosssheet_fraction_marks_output_sheet():
    import openpyxl
    from app.engines.extraction.pipeline.template_map import _crosssheet_fraction
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A3"], ws["B3"], ws["C3"] = "Particulars", 2024, 2025
    ws["A4"], ws["B4"], ws["C4"] = "Total assets", "='BS1'!B72", "='BS1'!C72"   # cross-sheet pulls
    ws["A5"], ws["B5"], ws["C5"] = "Equity", "='BS3'!B10", "='BS3'!C10"
    assert _crosssheet_fraction(ws, {2: 2024, 3: 2025}, 3) == 1.0


def test_build_face_truth_picks_up_analysis_of_sofp():
    # End-to-end: the analysis-of-SoFP primary table feeds balance-sheet face truth.
    truth = build_face_truth([_bs("Analysis of Statement of Financial Position")])
    assert truth[("total_assets", 2025)][0] == 266748030.0
    assert ("total_equity_and_liabilities", 2025) in truth


def test_split_off_asset_section_with_grand_total_is_primary():
    # The Millat 2025 case: the assets half of the SoFP was typed `non_current_assets`
    # (a sub-type note) but carries the grand total `total_assets` -> primary.
    t = FinancialTable(
        statement_type=StatementType.non_current_assets, title="ASSETS - NON-CURRENT ASSETS",
        unit_scale="thousands",
        line_items=[
            LineItem(label="Total non-current assets", canonical_metric="non_current_assets",
                     canonical_category="balance_sheet",
                     values=[LineItemValue(year=2025, value=8014208.0)]),
            LineItem(label="Total assets", canonical_metric="total_assets",
                     canonical_category="balance_sheet",
                     values=[LineItemValue(year=2025, value=32988591.0)]),
        ])
    assert infer_table_role(t) == "primary"
    truth = build_face_truth([t])
    assert truth[("total_assets", 2025)][0] == 32988591.0
    assert truth[("non_current_assets", 2025)][0] == 8014208.0


def test_ppe_note_without_grand_total_stays_note():
    # A real PP&E note (no balance-sheet grand total) must NOT be promoted.
    t = FinancialTable(
        statement_type=StatementType.non_current_assets, title="PROPERTY, PLANT AND EQUIPMENT",
        unit_scale="thousands",
        line_items=[LineItem(label="Operating fixed assets",
                             canonical_metric="property_plant_equipment",
                             canonical_category="balance_sheet",
                             values=[LineItemValue(year=2025, value=775150.0)])])
    assert infer_table_role(t) == "note"


def test_note_with_only_subtotal_not_promoted():
    # A note carrying a `current_assets` SUBTOTAL (not a grand total) — often a wrong
    # value — stays a note so it can't pollute face truth.
    t = FinancialTable(
        statement_type=StatementType.current_assets, title="Long-term loans and advances",
        unit_scale="thousands",
        line_items=[LineItem(label="Total current assets", canonical_metric="current_assets",
                             canonical_category="balance_sheet",
                             values=[LineItemValue(year=2024, value=22517991.0)])])
    assert infer_table_role(t) == "note"


def _income(title, role_metrics, unit="thousands"):
    return FinancialTable(statement_type=StatementType.income_statement, title=title,
                          unit_scale=unit,
                          line_items=[LineItem(label=lbl, canonical_metric=cm,
                                               canonical_category="income_statement",
                                               values=[LineItemValue(year=y, value=val)])
                                      for (lbl, cm, y, val) in role_metrics])


def test_note_fills_year_with_no_primary_candidate():
    # Primary income statement covers 2022-2024; the oldest comparative 2021 survives
    # only in a disaggregation NOTE -> note fills it as a fallback.
    primary = FinancialTable(statement_type=StatementType.income_statement,
                             title="Statement of Profit or Loss",
                             line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                                 canonical_category="income_statement",
                                 values=[LineItemValue(year=2022, value=53000.0),
                                         LineItemValue(year=2023, value=44000.0),
                                         LineItemValue(year=2024, value=91000.0)])])
    note = FinancialTable(statement_type=StatementType.revenue,   # sub-type -> note role
                          title="Revenue from contracts with customers",
                          line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                              canonical_category="income_statement",
                              values=[LineItemValue(year=2021, value=44930.0)])])
    truth = build_face_truth([primary, note])
    assert truth[("revenue", 2021)][0] == 44930.0      # filled from the note
    assert truth[("revenue", 2024)][0] == 91000.0      # primary unaffected


def test_primary_beats_note_for_same_year():
    primary = FinancialTable(statement_type=StatementType.income_statement,
                             title="Statement of Profit or Loss",
                             line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                                 canonical_category="income_statement",
                                 values=[LineItemValue(year=2024, value=91000.0)])])
    note = FinancialTable(statement_type=StatementType.revenue, title="Revenue note",
                          line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                              canonical_category="income_statement",
                              values=[LineItemValue(year=2024, value=88000.0)])])
    truth = build_face_truth([primary, note])
    assert truth[("revenue", 2024)][0] == 91000.0      # primary wins, note ignored


def test_analytical_never_supplies_truth_even_as_fallback():
    # gross_profit exists ONLY in a six-year summary -> must NOT become truth.
    summary = FinancialTable(statement_type=StatementType.financial_highlights,
                             title="Six Years at a Glance", unit_scale="thousands",
                             line_items=[LineItem(label="Gross profit", canonical_metric="gross_profit",
                                 canonical_category="income_statement",
                                 values=[LineItemValue(year=2021, value=9271.0)])])
    truth = build_face_truth([summary])
    assert ("gross_profit", 2021) not in truth


def test_share_capital_and_reserves_aliases_to_equity():
    # "Share capital and reserves" is the total-equity line -> supplies `equity` truth
    # (closes the years where the line is named that way rather than "Total equity").
    t = FinancialTable(
        statement_type=StatementType.balance_sheet,
        title="Statement of Financial Position", unit_scale="thousands",
        line_items=[
            LineItem(label="Total assets", canonical_metric="total_assets",
                     canonical_category="balance_sheet",
                     values=[LineItemValue(year=2022, value=184962368.0)]),
            LineItem(label="Share Capital & Reserves",
                     canonical_metric="share_capital_and_reserves",
                     canonical_category="balance_sheet",
                     values=[LineItemValue(year=2022, value=128540324.0)]),
        ])
    truth = build_face_truth([t])
    assert truth[("equity", 2022)][0] == 128540324.0     # aliased into `equity`
