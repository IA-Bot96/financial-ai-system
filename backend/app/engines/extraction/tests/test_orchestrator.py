"""Tests for Layer 7 orchestration (assembly + output routing)."""
import json
from pathlib import Path

import openpyxl
import pytest

from app.core.config import get_settings
from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.models.insight import Insight
from app.engines.extraction.models.result import DocumentResult
from app.engines.extraction.pipeline.orchestrator import process_documents


@pytest.fixture(autouse=True)
def _no_embeddings(monkeypatch):
    get_settings.cache_clear()
    monkeypatch.setenv("USE_EMBEDDINGS", "false")
    yield
    get_settings.cache_clear()


def _doc(report_year, st, lines) -> DocumentResult:
    table = FinancialTable(statement_type=st, title=st.value.replace("_", " ").title(), currency="PKR",
                           unit_scale="thousands", years=[report_year - 1, report_year], line_items=lines)
    ins = [Insight(year=report_year, source_report_year=report_year, area="x",
                   takeaway=f"point {report_year}", source_section="CEO Review", page=1, confidence=0.9)]
    return DocumentResult(file_name=f"r{report_year}.pdf", report_year=report_year, tables=[table], insights=ins)


def _rev_lines(y0, y1, local, export):
    return [
        LineItem(label="Local sales", values=[LineItemValue(year=y0, value=local[0]), LineItemValue(year=y1, value=local[1])]),
        LineItem(label="Export sales", values=[LineItemValue(year=y0, value=export[0]), LineItemValue(year=y1, value=export[1])]),
    ]


def test_no_template_writes_styled_workbook(tmp_path):
    results = [_doc(2025, StatementType.income_statement,
                    [LineItem(label="Revenue", values=[LineItemValue(year=2024, value=100.0), LineItemValue(year=2025, value=120.0)])])]
    out = tmp_path / "out.xlsx"
    res = process_documents(results, out)
    assert res.mode == "no_template" and res.plan is None
    # #8 observability: result carries a validation verdict, and a manifest is written.
    assert isinstance(res.production_ready, bool)
    assert res.manifest_path and Path(res.manifest_path).exists()
    manifest = json.loads(Path(res.manifest_path).read_text(encoding="utf-8"))
    assert manifest["production_ready"] == res.production_ready
    wb = openpyxl.load_workbook(out)
    assert "Income Statement" in wb.sheetnames and "Insights" in wb.sheetnames
    wb.close()


def _assert_seeded_history(path):
    wb = openpyxl.load_workbook(path)
    try:
        assert "Edit History" in wb.sheetnames                       # not "History" (ExcelJS-reserved)
        h = wb["Edit History"]
        assert [h.cell(1, c).value for c in range(1, 7)] == \
            ["Timestamp", "Sheet", "Cell", "Old", "New", "Saved"]   # exact header contract
        assert h.max_row == 1                                        # header only, no data rows
    finally:
        wb.close()


def test_history_sheet_seeded_both_paths(tmp_path):
    results = [_doc(2025, StatementType.income_statement,
                    [LineItem(label="Revenue", values=[LineItemValue(year=2024, value=100.0),
                                                        LineItemValue(year=2025, value=120.0)])])]
    # no-template path
    process_documents(results, tmp_path / "nt.xlsx")
    _assert_seeded_history(tmp_path / "nt.xlsx")
    # template path
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    process_documents([_doc(2025, StatementType.revenue, _rev_lines(2024, 2025, (50.0, 60.0), (30.0, 40.0)))],
                      tmp_path / "t.xlsx", template_path=tpl)
    _assert_seeded_history(tmp_path / "t.xlsx")


def test_history_append_is_idempotent(tmp_path):
    # Simulates re-extraction over a workbook whose History the app already wrote to.
    from app.engines.extraction.pipeline.excel_writer import append_insights_sheets, write_history_sheet
    p = tmp_path / "edited.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "PL1 - Revenue"
    wb["PL1 - Revenue"]["A1"] = "x"
    write_history_sheet(wb.create_sheet("Edit History"))
    wb["Edit History"].append(["2026-01-01T00:00", "PL1 - Revenue", "B2", "1", "2", "yes"])  # user edit
    wb.save(p)

    append_insights_sheets(p, [], [])     # re-run meta-sheet step
    wb2 = openpyxl.load_workbook(p)
    try:
        h = wb2["Edit History"]
        assert h.max_row == 2 and h.cell(2, 3).value == "B2"   # existing row preserved, not wiped
    finally:
        wb2.close()


def test_history_not_ingested_as_financial_data(tmp_path):
    from app.engines.fie.ingest.classify import classify_sheet
    from app.engines.fie.store import FinancialFactStore
    assert classify_sheet("Edit History") == "history"
    results = [_doc(2025, StatementType.income_statement,
                    [LineItem(label="Revenue", values=[LineItemValue(year=2024, value=100.0),
                                                        LineItemValue(year=2025, value=120.0)])])]
    process_documents(results, tmp_path / "wb.xlsx")
    store = FinancialFactStore.from_workbook(str(tmp_path / "wb.xlsx"))
    assert store.history == []                                  # empty log, not polluting data


def test_validation_review_toggle_gates_ledger_sheet(tmp_path, monkeypatch):
    results = [_doc(2025, StatementType.income_statement,
                    [LineItem(label="Revenue", values=[LineItemValue(year=2024, value=100.0),
                                                        LineItemValue(year=2025, value=120.0)])])]
    # Default ON -> review surface present.
    res_on = process_documents(results, tmp_path / "on.xlsx")
    wb = openpyxl.load_workbook(tmp_path / "on.xlsx")
    assert "Validation Ledger" in wb.sheetnames
    wb.close()
    assert json.loads(Path(res_on.manifest_path).read_text("utf-8"))["validation_review_enabled"] is True

    # Toggle OFF -> no review sheet, but the workbook + validation verdict still produced.
    monkeypatch.setenv("VALIDATION_REVIEW_ENABLED", "false")
    get_settings.cache_clear()
    res_off = process_documents(results, tmp_path / "off.xlsx")
    wb2 = openpyxl.load_workbook(tmp_path / "off.xlsx")
    assert "Validation Ledger" not in wb2.sheetnames     # review surface suppressed
    assert "Income Statement" in wb2.sheetnames           # workbook still produced
    wb2.close()
    man_off = json.loads(Path(res_off.manifest_path).read_text("utf-8"))
    assert man_off["validation_review_enabled"] is False
    assert isinstance(res_off.production_ready, bool)      # validation still computed


def test_review_off_keeps_corrections_and_provenance(tmp_path, monkeypatch):
    # Review OFF must NOT disable corrections (filled values) or provenance (Source Ledger).
    monkeypatch.setenv("VALIDATION_REVIEW_ENABLED", "false")
    get_settings.cache_clear()
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    results = [_doc(2025, StatementType.revenue, _rev_lines(2024, 2025, (50.0, 60.0), (30.0, 40.0)))]
    res = process_documents(results, tmp_path / "filled.xlsx", template_path=tpl)
    wb = openpyxl.load_workbook(tmp_path / "filled.xlsx")
    assert "Validation Ledger" not in wb.sheetnames      # review surface off
    assert "Source Ledger" in wb.sheetnames               # provenance kept (not a review artifact)
    assert wb["PL1 - Revenue"]["B5"].value == 50.0        # values still filled (correctness intact)
    wb.close()
    assert res.manifest_path and Path(res.manifest_path).exists()


def _make_template(path):
    wb = openpyxl.Workbook()
    pl = wb.active
    pl.title = "PL1 - Revenue"
    pl["A1"] = "Note 27 - Gross Revenue"
    pl["D2"] = "Forecasted"
    pl["A3"], pl["B3"], pl["C3"], pl["D3"] = "Particulars", 2024, 2025, 2026
    pl["A4"] = "GROSS REVENUE"
    pl["A5"] = "Local sales"
    pl["A6"] = "Export sales"
    pl["A7"], pl["B7"], pl["C7"] = "Gross Revenue", "=SUM(B5:B6)", "=SUM(C5:C6)"
    wb.save(path)


def test_template_fills_and_appends_insights(tmp_path):
    tpl = tmp_path / "tpl.xlsx"
    _make_template(tpl)
    results = [_doc(2025, StatementType.revenue, _rev_lines(2024, 2025, (50.0, 60.0), (30.0, 40.0)))]
    out = tmp_path / "filled.xlsx"

    res = process_documents(results, out, template_path=tpl)
    assert res.mode == "template" and res.plan and res.plan.writes

    wb = openpyxl.load_workbook(out)
    pl = wb["PL1 - Revenue"]
    assert pl["B5"].value == 50.0 and pl["C5"].value == 60.0   # Local sales filled
    assert pl["B7"].value == "=SUM(B5:B6)"                     # subtotal formula preserved
    assert "Insights" in wb.sheetnames                          # insights appended
    # #9 traceability: a Source Ledger sheet maps writes to their origin.
    assert "Source Ledger" in wb.sheetnames
    # #8 observability: manifest written alongside the workbook.
    assert res.manifest_path and Path(res.manifest_path).exists()
    wb.close()
