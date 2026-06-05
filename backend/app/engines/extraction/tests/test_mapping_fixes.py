"""Regression tests for the mapping-correctness fixes:
A) section-aware matching (no Local/Export duplication),
B) year-header rows rejected as values,
C) prefer unconsolidated over consolidated.
"""
import openpyxl
import pytest

from app.core.config import get_settings
from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.models.result import DocumentResult
from app.engines.extraction.models.table import RawTable
from app.engines.extraction.pipeline.multiyear import resolve_multiyear
from app.engines.extraction.pipeline.structure import build_financial_table
from app.engines.extraction.pipeline.template_map import build_plan


@pytest.fixture(autouse=True)
def _no_embeddings(monkeypatch):
    get_settings.cache_clear()
    monkeypatch.setenv("USE_EMBEDDINGS", "false")
    yield
    get_settings.cache_clear()


# --- B: year-header rejection + section tracking ---

def test_build_rejects_year_header_values_and_tracks_section():
    raw = RawTable(
        table_id="t", statement_type=StatementType.revenue, title="Revenue",
        header=["", "2024", "2025"], years=[2024, 2025],
        rows=[
            ["LOCAL SALES", "", ""],                       # section header (no values)
            ["Components consumed (Note 33.1)", "2024", "2025"],  # year header leaked in
            ["Tractors", "50", "60"],                      # real leaf
        ],
    )
    ft = build_financial_table(raw)
    labels = {li.label: li for li in ft.line_items}
    # The year-header row produced no values -> not emitted as a data line.
    assert "Components consumed (Note 33.1)" not in labels
    # Real leaf kept, tagged with its section.
    assert "Tractors" in labels
    tractors = labels["Tractors"]
    assert tractors.section == "LOCAL SALES"
    assert {v.year: v.value for v in tractors.values} == {2024: 50.0, 2025: 60.0}


# --- C: prefer unconsolidated ---

def test_multiyear_keeps_both_sets_labeled():
    # No template: BOTH unconsolidated and consolidated are preserved & labeled.
    res = DocumentResult(
        file_name="r2025.pdf", report_year=2025,
        tables=[
            FinancialTable(statement_type=StatementType.revenue, title="Revenue", consolidated=True,
                           line_items=[LineItem(label="Tractors", values=[LineItemValue(year=2025, value=50000.0)])]),
            FinancialTable(statement_type=StatementType.revenue, title="Revenue", consolidated=False,
                           line_items=[LineItem(label="Tractors", values=[LineItemValue(year=2025, value=5000.0)])]),
        ],
    )
    company = resolve_multiyear([res])
    by_flag = {t.consolidated: t for t in company.tables}
    assert set(by_flag) == {True, False}  # two separate labeled tables
    assert by_flag[False].line_items[0].values[0].value == 5000.0
    assert by_flag[True].line_items[0].values[0].value == 50000.0


def test_template_prefers_unconsolidated_candidates():
    import openpyxl
    from app.engines.extraction.pipeline.template_map import build_plan

    company = CompanyResult(company="Acme", fiscal_years=[2025], tables=[
        FinancialTable(statement_type=StatementType.revenue, title="Revenue", consolidated=False,
                       line_items=[LineItem(label="Net revenue", values=[LineItemValue(year=2025, value=5000.0)])]),
        FinancialTable(statement_type=StatementType.revenue, title="Revenue", consolidated=True,
                       line_items=[LineItem(label="Net revenue", values=[LineItemValue(year=2025, value=50000.0)])]),
    ])
    import tempfile, pathlib
    tpl = pathlib.Path(tempfile.mkdtemp()) / "t.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PL1 - Revenue"
    ws["A1"] = "Gross Revenue"
    ws["A3"], ws["B3"], ws["C3"] = "Particulars", 2024, 2025
    ws["A4"] = "Net revenue"
    wb.save(tpl)

    plan = build_plan(company, tpl)
    c4 = next((w for w in plan.writes if w.coordinate == "C4"), None)  # 2025 column
    assert c4 and c4.value == 5000.0  # unconsolidated chosen, not consolidated 50000


# --- A: section-aware template matching ---

def _company_local_only() -> CompanyResult:
    # Source has a LOCAL section only (no export breakdown).
    t = FinancialTable(
        statement_type=StatementType.revenue, title="Revenue",
        line_items=[
            LineItem(label="Tractors", section="LOCAL SALES",
                     values=[LineItemValue(year=2024, value=50.0), LineItemValue(year=2025, value=60.0)]),
        ],
    )
    return CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[t])


def _template(path):
    wb = openpyxl.Workbook()
    pl = wb.active
    pl.title = "PL1 - Revenue"
    pl["A1"] = "Gross Revenue"
    pl["A3"], pl["B3"], pl["C3"] = "Particulars", 2024, 2025
    pl["A4"] = "LOCAL SALES"
    pl["A5"] = "Tractors"      # should be filled from the local source line
    pl["A6"] = "EXPORT SALES"
    pl["A7"] = "Tractors"      # should stay BLANK (no export source)
    wb.save(path)


def test_section_aware_no_local_export_duplication(tmp_path):
    tpl = tmp_path / "t.xlsx"
    _template(tpl)
    plan = build_plan(_company_local_only(), tpl)

    by_coord = {(w.coordinate): w for w in plan.writes}
    # Local Tractors (row 5) filled...
    assert by_coord.get("B5") and by_coord["B5"].value == 50.0
    assert by_coord["B5"].template_label == "Tractors"
    # ...Export Tractors (row 7) NOT filled (no export breakdown in source).
    assert "B7" not in by_coord and "C7" not in by_coord
