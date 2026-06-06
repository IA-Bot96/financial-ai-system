"""Tests for reconciliation-row gap-filling on breakdown subtotals."""
import openpyxl

from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.services.face_truth import tieout
from app.engines.extraction.services.formula_eval import evaluate
from app.engines.extraction.services.validation import reconcile_breakdown_subtotals


def _company(nca_2025=8014208.0, nca_2024=8413955.0):
    return CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.balance_sheet,
                       title="Statement of Financial Position",
                       line_items=[LineItem(label="Total non-current assets",
                           canonical_metric="non_current_assets", canonical_category="balance_sheet",
                           values=[LineItemValue(year=2024, value=nca_2024),
                                   LineItemValue(year=2025, value=nca_2025)])])])


def _bs1(tmp_path, leaf_2025, leaf_2024=8413955):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BS1 - Non-Current Assets"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A4"], ws["F4"], ws["G4"] = "Property, plant and equipment", leaf_2024, leaf_2025
    ws["A6"], ws["F6"], ws["G6"] = "Total non-current assets", "=SUM(F4:F5)", "=SUM(G4:G5)"
    p = tmp_path / "wb.xlsx"; wb.save(p); return p


def test_material_gap_not_plugged(tmp_path):
    # Leaves 7,000,000 vs audited 8,014,208 -> 12.7% gap (MATERIAL) -> NOT plugged;
    # disclosed as DETAIL_INCOMPLETE (no fake precision).
    p = _bs1(tmp_path, leaf_2025=7000000)
    before = openpyxl.load_workbook(p)["BS1 - Non-Current Assets"]["G6"].value
    rows, n = reconcile_breakdown_subtotals(p, _company(), tieout, output_sheets=set())
    after = openpyxl.load_workbook(p)["BS1 - Non-Current Assets"]["G6"].value
    assert n == 0 and after == before                                            # not plugged
    assert any(r.status == "DETAIL_INCOMPLETE" and r.year == 2025 for r in rows)


def test_minor_gap_plugged(tmp_path):
    # Leaves 7,900,000 vs audited 8,014,208 -> 1.4% gap (<5%) -> DETAIL_PLUG (schedule foots).
    p = _bs1(tmp_path, leaf_2025=7900000)
    rows, n = reconcile_breakdown_subtotals(p, _company(), tieout, output_sheets=set())
    wb = openpyxl.load_workbook(p)
    assert abs(evaluate(wb, "BS1 - Non-Current Assets", "G6") - 8014208.0) <= 1   # now ties
    assert n == 1 and any(r.status == "DETAIL_PLUG" and r.year == 2025 for r in rows)
    assert wb["BS1 - Non-Current Assets"]["G6"].comment is not None               # documented


def test_subtotal_already_reconciles_not_touched(tmp_path):
    # Leaves already sum to the audited total -> no plug, formula unchanged.
    p = _bs1(tmp_path, leaf_2025=8014208)
    before = openpyxl.load_workbook(p)["BS1 - Non-Current Assets"]["G6"].value
    rows, n = reconcile_breakdown_subtotals(p, _company(), tieout, output_sheets=set())
    after = openpyxl.load_workbook(p)["BS1 - Non-Current Assets"]["G6"].value
    assert n == 0 and after == before


def test_section_subtotal_minor_gap_plugged(tmp_path):
    # A SECTION subtotal (non-key canonical) with a MINOR gap (<5%) is plugged; a section
    # is still reconciled when its leaves nearly foot. Leaves 4,050,000 vs LTI 4,173,730 = 3%.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BS1 - Non-Current Assets"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A10"], ws["F10"], ws["G10"] = "Investment in associates", 4000000, 4050000
    ws["A12"], ws["F12"], ws["G12"] = "Total Long-term Investments", "=SUM(F10:F11)", "=SUM(G10:G11)"
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.balance_sheet, title="Statement of Financial Position",
            line_items=[LineItem(label="Long-term investments", canonical_metric="long_term_investments",
                canonical_category="balance_sheet",
                values=[LineItemValue(year=2024, value=4000000.0), LineItemValue(year=2025, value=4173730.0)])])])
    rows, n = reconcile_breakdown_subtotals(p, company, tieout, output_sheets=set())
    wb2 = openpyxl.load_workbook(p)
    assert abs(evaluate(wb2, "BS1 - Non-Current Assets", "G12") - 4173730.0) <= 1
    assert any(r.metric == "long_term_investments" and r.status == "DETAIL_PLUG" for r in rows)


def test_section_subtotal_material_gap_disclosed(tmp_path):
    # Same section but leaves only 1,000,000 vs 4,173,730 (76% gap, MATERIAL) -> NOT plugged.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BS1 - Non-Current Assets"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A10"], ws["F10"], ws["G10"] = "Investment in associates", 900000, 1000000
    ws["A12"], ws["F12"], ws["G12"] = "Total Long-term Investments", "=SUM(F10:F11)", "=SUM(G10:G11)"
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.balance_sheet, title="Statement of Financial Position",
            line_items=[LineItem(label="Long-term investments", canonical_metric="long_term_investments",
                canonical_category="balance_sheet",
                values=[LineItemValue(year=2024, value=4000000.0), LineItemValue(year=2025, value=4173730.0)])])])
    rows, n = reconcile_breakdown_subtotals(p, company, tieout, output_sheets=set())
    wb2 = openpyxl.load_workbook(p)
    assert evaluate(wb2, "BS1 - Non-Current Assets", "G12") == 1000000        # unchanged, honest
    assert any(r.metric == "long_term_investments" and r.status == "DETAIL_INCOMPLETE" for r in rows)


def test_output_sheets_are_skipped(tmp_path):
    # If BS1 were (wrongly) declared an output sheet, the breakdown pass skips it.
    p = _bs1(tmp_path, leaf_2025=7000000)
    rows, n = reconcile_breakdown_subtotals(p, _company(), tieout,
                                            output_sheets={"BS1 - Non-Current Assets"})
    assert n == 0 and rows == []
