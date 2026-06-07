"""Defensive guards in the template-fill path:
  A — apply_plan never overwrites a formula cell (plan/template drift backstop).
  F — apply_plan refuses to write back over the template itself (self-overwrite).
  B — _wb_fingerprint counts formulas/styled cells (LibreOffice fidelity check).
"""
import openpyxl
import pytest

from app.engines.extraction.models.mapping import CellWrite, MappingPlan
from app.engines.extraction.pipeline.template_map import apply_plan
from app.engines.extraction.services.validation import _wb_fingerprint


def _template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BS"
    ws["A1"] = "Cash"
    ws["B1"] = None                 # empty input cell -> writable
    ws["B2"] = "=B1*2"              # formula cell -> must never be clobbered
    wb.save(path)
    return path


def test_apply_plan_skips_formula_cells(tmp_path):
    tpl = _template(tmp_path / "tpl.xlsx")
    out = tmp_path / "out.xlsx"
    plan = MappingPlan(writes=[
        CellWrite(sheet="BS", coordinate="B1", year=2024, value=100.0, template_label="Cash"),
        CellWrite(sheet="BS", coordinate="B2", year=2024, value=999.0, template_label="x"),  # targets a formula
    ])
    apply_plan(plan, tpl, out)

    wb = openpyxl.load_workbook(out, data_only=False)
    ws = wb["BS"]
    assert ws["B1"].value == 100.0            # empty cell written
    assert ws["B2"].value == "=B1*2"          # formula preserved, NOT replaced by 999


def test_apply_plan_refuses_self_overwrite(tmp_path):
    tpl = _template(tmp_path / "tpl.xlsx")
    plan = MappingPlan(writes=[])
    with pytest.raises(ValueError, match="same file"):
        apply_plan(plan, tpl, tpl)            # template == output


def test_wb_fingerprint_counts_formulas_and_styles(tmp_path):
    p = tmp_path / "fp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=1+1"
    ws["A2"] = "=A1*3"
    ws["A3"] = 5
    from openpyxl.styles import Font
    ws["A3"].font = Font(bold=True)           # one explicitly styled cell
    wb.save(p)

    formulas, styled = _wb_fingerprint(p)
    assert formulas == 2
    assert styled >= 1                        # at least the bold cell carries a style
