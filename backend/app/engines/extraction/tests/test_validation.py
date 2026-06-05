"""Tests for the validation primitives: formula evaluator, computed tie-out,
scoped resolution (P2), and merge source-ranking (#11)."""
import openpyxl

from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.models.result import DocumentResult
from app.engines.extraction.pipeline.gpt_tables import _resolve_canonicals
from app.engines.extraction.pipeline.multiyear import resolve_multiyear
from app.engines.extraction.services.face_truth import build_face_truth, tieout
from app.engines.extraction.services.formula_eval import evaluate
from app.engines.extraction.services.metric_resolver import get_resolver


def test_formula_eval_grammar():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws["B4"], ws["B5"] = 10, 20
    ws["B6"], ws["B7"], ws["B8"] = "=SUM(B4:B5)", "=B6+B4-B5", "=$B$4+$B$5"
    rev = wb.create_sheet("PL1 - Revenue"); rev["B29"] = 999
    ws["B9"], ws["B10"] = "='PL1 - Revenue'!B29", "=-'PL1 - Revenue'!B29"
    ws["B11"] = "=B4*B5"               # unsupported operator
    assert evaluate(wb, "S", "B6") == 30.0
    assert evaluate(wb, "S", "B7") == 20.0
    assert evaluate(wb, "S", "B8") == 30.0
    assert evaluate(wb, "S", "B9") == 999.0
    assert evaluate(wb, "S", "B10") == -999.0
    assert evaluate(wb, "S", "B11") is None        # bails, never false-evaluates


def test_formula_eval_circular_is_safe():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws["A1"], ws["A2"] = "=A2", "=A1"
    assert evaluate(wb, "S", "A1") is None          # circular -> None, no crash


def test_scoped_resolution_demotes_cross_family_metric():
    # 'Cost of sales' (income metric) inside a PP&E (balance) note -> demoted to None,
    # so it can't pollute truth or be mis-mapped as the headline metric.
    t = FinancialTable(statement_type=StatementType.non_current_assets, title="PP&E",
                       line_items=[LineItem(label="Cost of sales",
                                            values=[LineItemValue(year=2025, value=34576.0)])])
    _resolve_canonicals(t, get_resolver())
    li = t.line_items[0]
    assert li.canonical_metric is None and li.resolution == "no_confident_metric"


def test_merge_prefers_primary_over_note():
    # Same metric/year in a primary face table (1000) and a note table (9999) -> primary wins.
    primary = FinancialTable(statement_type=StatementType.income_statement,
                             title="Statement of Profit or Loss",
                             line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                                 canonical_category="income_statement",
                                 values=[LineItemValue(year=2025, value=1000.0)])])
    note = FinancialTable(statement_type=StatementType.income_statement, title="Revenue note",
                          line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                              canonical_category="income_statement",
                              values=[LineItemValue(year=2025, value=9999.0)])])
    res = DocumentResult(file_name="r2025.pdf", report_year=2025, tables=[primary, note])
    company = resolve_multiyear([res])
    face = build_face_truth(company.tables)
    assert face[("revenue", 2025)][0] == 1000.0     # primary, not the note's 9999


def test_p1_rejects_extra_digit_outlier_and_picks_consistent_value():
    # An extra-digit (10x) OCR error must not become truth; the consistent value wins.
    def tbl(val):
        return FinancialTable(statement_type=StatementType.balance_sheet,
                              title="Statement of Financial Position",
                              line_items=[LineItem(label="Total assets", canonical_metric="total_assets",
                                  canonical_category="balance_sheet",
                                  values=[LineItemValue(year=2025, value=val)])])
    face = build_face_truth([tbl(100.0), tbl(100.0), tbl(1000.0)])  # 1000 = 10x slip
    assert face[("total_assets", 2025)][0] == 100.0  # outlier rejected, consistent value chosen


def test_computed_ledger_sign_scoping(tmp_path):
    # A breakdown sheet's positive-magnitude cost formula must NOT be flagged against
    # a negative face value on sign alone (magnitude scoping); the same on an OUTPUT
    # sheet IS sign-sensitive.
    import openpyxl
    from app.engines.extraction.services.validation import computed_output_ledger

    wb = openpyxl.Workbook()
    pl2 = wb.active; pl2.title = "PL2 - Cost of Sales"
    pl2["A3"], pl2["B3"], pl2["C3"] = "Particulars", 2024, 2025   # >=2 year columns
    pl2["A4"], pl2["B4"], pl2["C4"] = "Raw material", 60, 60
    pl2["A5"], pl2["B5"], pl2["C5"] = "Labour", 40, 40
    pl2["A6"], pl2["B6"], pl2["C6"] = "Cost of sales", "=SUM(B4:B5)", "=SUM(C4:C5)"  # = +100 (note)
    out = wb.create_sheet("P&L")
    out["A3"], out["B3"], out["C3"] = "Particulars", 2024, 2025
    out["A4"], out["B4"], out["C4"] = "Cost of sales", \
        "=-'PL2 - Cost of Sales'!B6", "=-'PL2 - Cost of Sales'!C6"  # = -100 (output)
    p = tmp_path / "wb.xlsx"; wb.save(p)

    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label="Cost of sales", canonical_metric="cost_of_sales",
                           canonical_category="income_statement",
                           values=[LineItemValue(year=2024, value=-100.0),
                                   LineItemValue(year=2025, value=-100.0)])])])  # face: negative
    # Sign sensitivity is per-formula: PL2 (intra-sheet SUM) uses magnitude; P&L
    # (cross-sheet pull) is signed — no reliance on a sheet set.
    rows, fails, _un = computed_output_ledger(p, company, tieout)
    pl2_rows = [r for r in rows if r.sheet.startswith("PL2")]
    out_rows = [r for r in rows if r.sheet == "P&L"]
    assert pl2_rows and all(r.status == "ok" for r in pl2_rows)   # +100 vs |-100| -> ok (magnitude)
    assert out_rows and all(r.status == "ok" for r in out_rows)   # -100 vs -100 -> ok (signed)


def test_computed_ledger_signed_catches_cross_sheet_sign_error(tmp_path):
    # A cross-sheet output pull with the WRONG sign IS caught (signed comparison).
    import openpyxl
    from app.engines.extraction.services.validation import computed_output_ledger
    wb = openpyxl.Workbook()
    pl2 = wb.active; pl2.title = "PL2 - Cost of Sales"
    pl2["A3"], pl2["B3"], pl2["C3"] = "Particulars", 2024, 2025
    pl2["A6"], pl2["B6"], pl2["C6"] = "Cost of sales", 100, 100
    out = wb.create_sheet("P&L")
    out["A3"], out["B3"], out["C3"] = "Particulars", 2024, 2025
    # BUG: forgot the minus on the pull -> +100 where face is -100.
    out["A4"], out["B4"], out["C4"] = "Cost of sales", \
        "='PL2 - Cost of Sales'!B6", "='PL2 - Cost of Sales'!C6"
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label="Cost of sales", canonical_metric="cost_of_sales",
                           canonical_category="income_statement",
                           values=[LineItemValue(year=2024, value=-100.0),
                                   LineItemValue(year=2025, value=-100.0)])])])
    rows, fails, _un = computed_output_ledger(p, company, tieout)
    assert fails >= 1 and any(r.sheet == "P&L" and r.status == "MISMATCH" for r in rows)


def test_unevaluable_key_formula_is_surfaced_not_passed(tmp_path):
    # A key-metric formula the evaluator can't parse must be SURFACED (UNEVALUATED +
    # counted + commented), never silently treated as a pass.
    import openpyxl
    from app.engines.extraction.services.validation import computed_output_ledger
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "P&L"
    ws["A3"], ws["B3"], ws["C3"] = "Particulars", 2024, 2025
    ws["A4"], ws["B4"], ws["C4"] = "Revenue", "=VLOOKUP(B1,X,2)", "=B5*C5"  # unsupported
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                           canonical_category="income_statement",
                           values=[LineItemValue(year=2024, value=100.0),
                                   LineItemValue(year=2025, value=100.0)])])])
    rows, fails, unevaluable = computed_output_ledger(p, company, tieout, output_sheets={"P&L"})
    assert fails == 0 and unevaluable == 2                      # not failures, but surfaced
    assert all(r.status == "UNEVALUATED" for r in rows)
    assert openpyxl.load_workbook(p)["P&L"]["B4"].comment is not None  # in-cell flag


def test_same_sheet_output_formula_is_signed_when_output_sheet(tmp_path):
    # A same-sheet output formula (no cross-sheet '!') with the wrong sign must be
    # caught when its sheet is a declared output sheet (signed), not magnitude-only.
    import openpyxl
    from app.engines.extraction.services.validation import computed_output_ledger
    wb = openpyxl.Workbook()
    pl = wb.active; pl.title = "P&L"
    pl["A3"], pl["B3"], pl["C3"] = "Particulars", 2024, 2025
    pl["A4"], pl["B4"], pl["C4"] = "Revenue", 300, 300
    pl["A5"], pl["B5"], pl["C5"] = "Cost", -200, -200
    # BUG: gross profit computed as -(rev+cost) = -100 (same-sheet, no '!').
    pl["A6"], pl["B6"], pl["C6"] = "Gross profit", "=-(B4+B5)", "=-(C4+C5)"
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label="Gross profit", canonical_metric="gross_profit",
                           canonical_category="income_statement",
                           values=[LineItemValue(year=2024, value=100.0),
                                   LineItemValue(year=2025, value=100.0)])])])
    # As an output sheet -> signed -> the -100 is caught.
    _rows, fails, _un = computed_output_ledger(p, company, tieout, output_sheets={"P&L"})
    assert fails >= 1
    # As a (hypothetical) breakdown sheet -> magnitude -> |-100| ties out, not flagged.
    _rows2, fails2, _un2 = computed_output_ledger(p, company, tieout, output_sheets=set())
    assert fails2 == 0
