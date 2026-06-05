"""Tests for Bucket-A template-author formula repair."""
import openpyxl

from app.engines.extraction.services.formula_repair import repair_template_formulas


def test_frozen_absolute_ref_is_rerelativized(tmp_path):
    # =$C$10+$C$17 dragged across C18:F18 (column locked) -> each column tracks itself.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PL1 - Revenue"
    ws["A3"] = "Particulars"
    ws["C3"], ws["D3"], ws["E3"], ws["F3"] = 2022, 2023, 2024, 2025
    for col in ("C", "D", "E", "F"):
        ws[f"{col}10"] = f"=SUM({col}5:{col}9)"
        ws[f"{col}17"] = f"=SUM({col}13:{col}16)"
        ws[f"{col}18"] = "=$C$10+$C$17"          # frozen on column C
    p = tmp_path / "wb.xlsx"; wb.save(p)

    repairs = repair_template_formulas(p)
    out = openpyxl.load_workbook(p)["PL1 - Revenue"]
    # C18 already tracks its own column (=$C$10+$C$17 IS correct in column C) -> left as-is.
    assert out["C18"].value == "=$C$10+$C$17"
    assert out["D18"].value == "=D10+D17"
    assert out["E18"].value == "=E10+E17"
    assert out["F18"].value == "=F10+F17"
    assert {r.coordinate for r in repairs if r.kind == "frozen-ref"} == {"D18", "E18", "F18"}


def test_frozen_ref_with_correct_leading_columns(tmp_path):
    # The real Millat shape: B18/C18 are correct (each self-references), only D:F were
    # drag-copied frozen on $C$. Repair D:F, leave B18/C18 untouched.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PL1 - Revenue"
    ws["A3"] = "Particulars"
    ws["B3"], ws["C3"], ws["D3"], ws["E3"], ws["F3"] = 2021, 2022, 2023, 2024, 2025
    ws["B18"] = "=$B$10+$B$17"
    ws["C18"] = "=$C$10+$C$17"
    ws["D18"] = ws["E18"] = ws["F18"] = "=$C$10+$C$17"   # frozen drag of C
    p = tmp_path / "wb.xlsx"; wb.save(p)

    repairs = repair_template_formulas(p)
    out = openpyxl.load_workbook(p)["PL1 - Revenue"]
    assert out["B18"].value == "=$B$10+$B$17"     # correct, untouched
    assert out["C18"].value == "=$C$10+$C$17"     # correct, untouched
    assert out["D18"].value == "=D10+D17"
    assert out["E18"].value == "=E10+E17"
    assert out["F18"].value == "=F10+F17"
    assert {r.coordinate for r in repairs if r.kind == "frozen-ref"} == {"D18", "E18", "F18"}


def test_hardcoded_zero_check_restored_from_sibling(tmp_path):
    # C58 holds the real =C57-C28 tie-out; D58:G58 were masked to literal 0.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Balance Sheet"
    ws["A2"] = "Particulars"
    ws["C2"], ws["D2"], ws["E2"], ws["F2"], ws["G2"] = 2021, 2022, 2023, 2024, 2025
    ws["C58"] = "=C57-C28"
    ws["D58"] = ws["E58"] = ws["F58"] = ws["G58"] = 0
    p = tmp_path / "wb.xlsx"; wb.save(p)

    repairs = repair_template_formulas(p)
    out = openpyxl.load_workbook(p)["Balance Sheet"]
    assert out["C58"].value == "=C57-C28"          # untouched
    assert out["D58"].value == "=D57-D28"
    assert out["G58"].value == "=G57-G28"
    assert {r.coordinate for r in repairs if r.kind == "constant-check"} == {"D58", "E58", "F58", "G58"}


def test_legitimately_fixed_ref_to_outside_band_is_untouched(tmp_path):
    # =$B$2 (an assumptions cell OUTSIDE the year band) repeated across columns is a
    # genuine constant, NOT the frozen defect -> must be left alone.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PL1 - Revenue"
    ws["A3"] = "Particulars"
    ws["C3"], ws["D3"], ws["E3"], ws["F3"] = 2022, 2023, 2024, 2025
    ws["B2"] = 0.17
    for col in ("C", "D", "E", "F"):
        ws[f"{col}19"] = "=$B$2"
    p = tmp_path / "wb.xlsx"; wb.save(p)

    repairs = repair_template_formulas(p)
    out = openpyxl.load_workbook(p)["PL1 - Revenue"]
    assert all(out[f"{c}19"].value == "=$B$2" for c in ("C", "D", "E", "F"))
    assert repairs == []


def test_pure_data_rows_and_lone_zeros_untouched(tmp_path):
    # A data row with a 0 but NO sibling formula must not be rewritten.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BS1 - Non-Current Assets"
    ws["A3"] = "Particulars"
    ws["C3"], ws["D3"], ws["E3"], ws["F3"] = 2022, 2023, 2024, 2025
    ws["C20"], ws["D20"], ws["E20"], ws["F20"] = 0, 5, 6, 7
    p = tmp_path / "wb.xlsx"; wb.save(p)

    repairs = repair_template_formulas(p)
    out = openpyxl.load_workbook(p)["BS1 - Non-Current Assets"]
    assert out["C20"].value == 0 and out["D20"].value == 5
    assert repairs == []


def test_relative_per_column_formulas_not_flagged_as_frozen(tmp_path):
    # Each column already references itself (=C5, =D5, ...) -> distinct, not frozen.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PL2 - Cost of Sales"
    ws["A3"] = "Particulars"
    ws["C3"], ws["D3"], ws["E3"], ws["F3"] = 2022, 2023, 2024, 2025
    for col in ("C", "D", "E", "F"):
        ws[f"{col}47"] = f"=SUM({col}40:{col}46)"
    p = tmp_path / "wb.xlsx"; wb.save(p)

    repairs = repair_template_formulas(p)
    out = openpyxl.load_workbook(p)["PL2 - Cost of Sales"]
    assert out["D47"].value == "=SUM(D40:D46)"     # unchanged
    assert repairs == []
