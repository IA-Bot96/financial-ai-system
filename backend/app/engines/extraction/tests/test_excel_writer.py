"""Tests for Layer 6 styled Excel writer (no-template path)."""
import openpyxl

from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.models.insight import Insight
from app.engines.extraction.pipeline.excel_writer import write_company_workbook
from app.engines.extraction.services import styles as S


def _company() -> CompanyResult:
    table = FinancialTable(
        statement_type=StatementType.income_statement, title="Income Statement",
        currency="PKR", unit_scale="thousands", years=[2024, 2025],
        line_items=[
            LineItem(label="Revenue from contracts with customers",
                     values=[LineItemValue(year=2024, value=95020.0), LineItemValue(year=2025, value=53347.0)]),
            LineItem(label="Cost of sales",
                     values=[LineItemValue(year=2024, value=-71048.0), LineItemValue(year=2025, value=-38940.0)]),
            LineItem(label="Gross profit",
                     values=[LineItemValue(year=2024, value=23972.0), LineItemValue(year=2025, value=14407.0)]),
        ],
    )
    ins = [Insight(year=2025, source_report_year=2025, area="Margins",
                   takeaway="Gross margin held.", source_section="CEO Review", page=10, confidence=0.91)]
    review = [Insight(year=2025, source_report_year=2025, area="Risk",
                      takeaway="Mixed signal.", source_section="Risks", page=20, confidence=0.6)]
    return CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[table],
                         insights=ins, insights_review=review)


def test_workbook_structure_and_values(tmp_path):
    out = tmp_path / "out.xlsx"
    write_company_workbook(_company(), out)
    wb = openpyxl.load_workbook(out)
    # History = app-managed change log, seeded (header-only) on every produced workbook.
    assert wb.sheetnames == ["Income Statement", "Insights", "Insights Review", "History"]
    assert [wb["History"].cell(1, c).value for c in range(1, 7)] == \
        ["Timestamp", "Sheet", "Cell", "Old", "New", "Saved"] and wb["History"].max_row == 1

    ws = wb["Income Statement"]
    assert ws["A1"].value == "Income Statement"
    assert ws["A2"].value == "(PKR in thousands)"
    assert ws["A3"].value == "Particulars" and ws["B3"].value == 2024 and ws["C3"].value == 2025
    assert ws["A4"].value == "Revenue from contracts with customers"
    assert ws["B4"].value == 95020.0 and ws["C4"].value == 53347.0
    assert ws["C5"].value == -38940.0
    assert ws.freeze_panes == "B4"
    wb.close()


def test_template_styling_applied(tmp_path):
    out = tmp_path / "out.xlsx"
    write_company_workbook(_company(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Income Statement"]

    # Navy header, white bold, parens number format.
    assert ws["A1"].fill.fgColor.rgb == S.NAVY and ws["A1"].font.color.rgb == S.WHITE
    assert ws["A3"].fill.fgColor.rgb == S.NAVY and ws["A3"].font.bold
    assert ws["A2"].fill.fgColor.rgb == S.UNIT_BLUE
    assert ws["B4"].number_format == S.NUMBER_FORMAT
    # "Gross profit" -> total row -> green + bold.
    assert ws["A6"].value == "Gross profit"
    assert ws["A6"].fill.fgColor.rgb == S.TOTAL_GREEN and ws["A6"].font.bold

    ins = wb["Insights"]
    assert [ins.cell(1, c).value for c in range(1, 8)] == [
        "Year", "Source Report Year", "Area", "Takeaway", "Source Section", "Page", "Confidence"]
    assert ins["A1"].fill.fgColor.rgb == S.NAVY
    assert ins["D2"].value == "Gross margin held." and ins["G2"].value == 0.91
    assert ins.freeze_panes == "A2"
    wb.close()


def _company_with_subtotals() -> CompanyResult:
    table = FinancialTable(
        statement_type=StatementType.revenue, title="Revenue", currency="PKR",
        unit_scale="thousands", years=[2024, 2025],
        line_items=[
            LineItem(label="Tractors", values=[LineItemValue(year=2024, value=40.0), LineItemValue(year=2025, value=50.0)]),
            LineItem(label="Implements", values=[LineItemValue(year=2024, value=20.0), LineItemValue(year=2025, value=30.0)]),
            LineItem(label="Gross Local Sales", values=[LineItemValue(year=2024, value=60.0), LineItemValue(year=2025, value=80.0)]),  # subtotal
            LineItem(label="Net Revenue", values=[LineItemValue(year=2024, value=60.0), LineItemValue(year=2025, value=80.0)]),       # grand total (block above is a total)
        ],
    )
    return CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[table])


def test_subtotal_rows_become_sum_formulas(tmp_path):
    out = tmp_path / "out.xlsx"
    write_company_workbook(_company_with_subtotals(), out)
    ws = openpyxl.load_workbook(out)["Revenue"]
    # rows: 4 Tractors, 5 Implements, 6 Gross Local Sales (subtotal), 7 Net Revenue (grand total)
    assert ws["B4"].value == 40.0 and ws["B5"].value == 20.0          # leaves untouched
    assert ws["B6"].value == "=SUM(B4:B5)" and ws["C6"].value == "=SUM(C4:C5)"  # subtotal -> formula
    # Net Revenue's block above is a total (no leaf run) -> reported value kept, NOT a formula.
    assert ws["B7"].value == 60.0


def test_two_sections_each_sum_their_own_leaves(tmp_path):
    table = FinancialTable(
        statement_type=StatementType.operating_expenses, title="Expenses", years=[2025],
        line_items=[
            LineItem(label="Dist Salaries", values=[LineItemValue(year=2025, value=10.0)]),
            LineItem(label="Dist Fuel", values=[LineItemValue(year=2025, value=5.0)]),
            LineItem(label="Total Distribution", values=[LineItemValue(year=2025, value=15.0)]),
            LineItem(label="Admin Salaries", values=[LineItemValue(year=2025, value=8.0)]),
            LineItem(label="Admin Fuel", values=[LineItemValue(year=2025, value=4.0)]),
            LineItem(label="Total Administrative", values=[LineItemValue(year=2025, value=12.0)]),
        ],
    )
    out = tmp_path / "e.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["Expenses"]
    # rows: 4,5 dist leaves; 6 Total Distribution; 7,8 admin leaves; 9 Total Administrative
    assert ws["B6"].value == "=SUM(B4:B5)"   # distribution sums its own leaves
    assert ws["B9"].value == "=SUM(B7:B8)"   # admin sums its own leaves (not distribution's)


def test_grand_total_adds_section_subtotals(tmp_path):
    # Type 2/3: a total whose block above is other subtotals becomes their addition.
    table = FinancialTable(
        statement_type=StatementType.operating_expenses, title="OpEx", years=[2025],
        line_items=[
            LineItem(label="Dist Salaries", values=[LineItemValue(year=2025, value=10.0)]),
            LineItem(label="Dist Fuel", values=[LineItemValue(year=2025, value=5.0)]),
            LineItem(label="Total Distribution", values=[LineItemValue(year=2025, value=15.0)]),
            LineItem(label="Admin Salaries", values=[LineItemValue(year=2025, value=8.0)]),
            LineItem(label="Admin Fuel", values=[LineItemValue(year=2025, value=4.0)]),
            LineItem(label="Total Administrative", values=[LineItemValue(year=2025, value=12.0)]),
            LineItem(label="Total Operating Expenses", values=[LineItemValue(year=2025, value=27.0)]),
        ],
    )
    out = tmp_path / "g.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["OpEx"]
    # rows: 6 Total Distribution, 9 Total Administrative, 10 Total Operating Expenses
    assert ws["B6"].value == "=SUM(B4:B5)"
    assert ws["B9"].value == "=SUM(B7:B8)"
    assert ws["B10"].value == "=B6+B9"      # grand total adds the two subtotals (15+12=27)


def test_tieout_mismatch_keeps_reported_value(tmp_path):
    # Running subtotal / wrong sign: computed sum != reported -> keep reported, no formula.
    table = FinancialTable(
        statement_type=StatementType.income_statement, title="IS", years=[2025],
        line_items=[
            LineItem(label="Revenue", values=[LineItemValue(year=2025, value=100.0)]),
            LineItem(label="Cost of sales", values=[LineItemValue(year=2025, value=60.0)]),  # positive (wrong sign for a blind SUM)
            LineItem(label="Gross profit", values=[LineItemValue(year=2025, value=40.0)]),   # 100-60, NOT 100+60
        ],
    )
    out = tmp_path / "t.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["IS"]
    assert ws["B6"].value == 40.0           # 100+60=160 != 40 -> reported value kept, no =SUM


def test_per_year_guard_keeps_reported_when_leaves_blank(tmp_path):
    # Leaves only have 2024; the subtotal has a reported 2025 -> 2025 stays reported, not SUM-of-blanks=0.
    table = FinancialTable(
        statement_type=StatementType.revenue, title="Rev", years=[2024, 2025],
        line_items=[
            LineItem(label="Tractors", values=[LineItemValue(year=2024, value=40.0)]),
            LineItem(label="Implements", values=[LineItemValue(year=2024, value=20.0)]),
            LineItem(label="Gross Local Sales",
                     values=[LineItemValue(year=2024, value=60.0), LineItemValue(year=2025, value=80.0)]),
        ],
    )
    out = tmp_path / "y.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["Rev"]
    assert ws["B6"].value == "=SUM(B4:B5)"  # 2024 leaves present -> formula
    assert ws["C6"].value == 80.0           # 2025 leaves blank -> reported value kept (NOT 0)


def test_role_overrides_total_heuristic_false_positive(tmp_path):
    # "Profit on bank deposits" would trip the is_total_row prefix heuristic, but
    # role="leaf" makes it a leaf -> it sums INTO the total below.
    table = FinancialTable(
        statement_type=StatementType.other_income, title="OI", years=[2025],
        line_items=[
            LineItem(label="Interest income", role="leaf", values=[LineItemValue(year=2025, value=10.0)]),
            LineItem(label="Profit on bank deposits", role="leaf", values=[LineItemValue(year=2025, value=5.0)]),
            LineItem(label="Total other income", role="total", values=[LineItemValue(year=2025, value=15.0)]),
        ],
    )
    out = tmp_path / "oi.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["OI"]
    assert ws["B6"].value == "=SUM(B4:B5)"   # both leaves summed (heuristic would have dropped row5)


def test_explicit_components_handle_running_subtotal(tmp_path):
    # Income statement: Gross profit = Revenue - Cost; Operating profit = Gross - Dist - Admin.
    # Positional grouping cannot express these; explicit components + is_contra can.
    table = FinancialTable(
        statement_type=StatementType.income_statement, title="P&L", years=[2025],
        line_items=[
            LineItem(label="Revenue", role="leaf", values=[LineItemValue(year=2025, value=100.0)]),
            LineItem(label="Cost of sales", role="leaf", is_contra=True, values=[LineItemValue(year=2025, value=60.0)]),
            LineItem(label="Gross profit", role="total", components=["Revenue", "Cost of sales"],
                     values=[LineItemValue(year=2025, value=40.0)]),
            LineItem(label="Distribution costs", role="leaf", is_contra=True, values=[LineItemValue(year=2025, value=5.0)]),
            LineItem(label="Administrative expenses", role="leaf", is_contra=True, values=[LineItemValue(year=2025, value=3.0)]),
            LineItem(label="Operating profit", role="total",
                     components=["Gross profit", "Distribution costs", "Administrative expenses"],
                     values=[LineItemValue(year=2025, value=32.0)]),
        ],
    )
    out = tmp_path / "pl.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["P&L"]
    # rows: 4 Revenue, 5 Cost, 6 Gross profit, 7 Dist, 8 Admin, 9 Operating profit
    assert ws["B6"].value == "=B4-B5"          # 100 - 60 = 40 (cost is contra)
    assert ws["B9"].value == "=B6-B7-B8"       # 40 - 5 - 3 = 32 (running subtotal)


def test_contra_leaf_in_positional_block_subtracts(tmp_path):
    # No components, but a contra leaf in the block -> signed reference list, not SUM.
    table = FinancialTable(
        statement_type=StatementType.revenue, title="NetSales", years=[2025],
        line_items=[
            LineItem(label="Gross sales", role="leaf", values=[LineItemValue(year=2025, value=100.0)]),
            LineItem(label="Trade discount", role="leaf", is_contra=True, values=[LineItemValue(year=2025, value=10.0)]),
            LineItem(label="Net sales", role="subtotal", values=[LineItemValue(year=2025, value=90.0)]),
        ],
    )
    out = tmp_path / "ns.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["NetSales"]
    assert ws["B6"].value == "=B4-B5"          # 100 - 10 = 90 (discount subtracted, not SUM)


def test_no_formula_filled_into_blank_total_cell(tmp_path):
    # A total with NO reported value for a year must stay blank, never be back-filled
    # with a (possibly mis-grouped) computed sum -- no reported value to validate against.
    table = FinancialTable(
        statement_type=StatementType.revenue, title="Rev2", years=[2024, 2025],
        line_items=[
            LineItem(label="Tractors",
                     values=[LineItemValue(year=2024, value=40.0), LineItemValue(year=2025, value=50.0)]),
            LineItem(label="Implements",
                     values=[LineItemValue(year=2024, value=20.0), LineItemValue(year=2025, value=30.0)]),
            LineItem(label="Gross Local Sales",
                     values=[LineItemValue(year=2024, value=60.0)]),  # no 2025 reported value
        ],
    )
    out = tmp_path / "b.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2024, 2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["Rev2"]
    assert ws["B6"].value == "=SUM(B4:B5)"  # 2024 reported (60) ties out -> formula
    assert ws["C6"].value is None           # 2025 total blank -> NOT back-filled with 80


def test_failed_grand_total_does_not_discard_subtotals(tmp_path):
    # A grand total whose reported value is garbled must not be emitted AND must not
    # discard the section subtotals a later grand total needs.
    table = FinancialTable(
        statement_type=StatementType.operating_expenses, title="OpEx2", years=[2025],
        line_items=[
            LineItem(label="Dist A", values=[LineItemValue(year=2025, value=10.0)]),
            LineItem(label="Dist B", values=[LineItemValue(year=2025, value=5.0)]),
            LineItem(label="Total Distribution", values=[LineItemValue(year=2025, value=15.0)]),
            LineItem(label="Admin A", values=[LineItemValue(year=2025, value=8.0)]),
            LineItem(label="Admin B", values=[LineItemValue(year=2025, value=4.0)]),
            LineItem(label="Total Administrative", values=[LineItemValue(year=2025, value=12.0)]),
            LineItem(label="Total Operating Expenses", values=[LineItemValue(year=2025, value=999.0)]),  # garbled
            LineItem(label="Total Charges", values=[LineItemValue(year=2025, value=27.0)]),  # = Dist + Admin
        ],
    )
    out = tmp_path / "d.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["OpEx2"]
    # rows: 6 Total Distribution, 9 Total Administrative, 10 Total Operating Expenses, 11 Total Charges
    assert ws["B10"].value == 999.0          # garbled grand total kept as reported, no formula
    assert ws["B11"].value == "=B6+B9"       # subtotals NOT discarded -> later total still sums them (15+12=27)


def test_small_magnitude_mismatch_is_rejected(tmp_path):
    # Millions-scale: reported 3.0 vs leaves summing to 3.1 (3% off) must NOT tie out;
    # the reported value is kept, not overwritten with a wrong computed sum.
    table = FinancialTable(
        statement_type=StatementType.revenue, title="Small", years=[2025],
        line_items=[
            LineItem(label="Line A", values=[LineItemValue(year=2025, value=1.2)]),
            LineItem(label="Line B", values=[LineItemValue(year=2025, value=1.9)]),
            LineItem(label="Total", values=[LineItemValue(year=2025, value=3.0)]),  # != 3.1
        ],
    )
    out = tmp_path / "s.xlsx"
    write_company_workbook(CompanyResult(company="Acme", fiscal_years=[2025], tables=[table]), out)
    ws = openpyxl.load_workbook(out)["Small"]
    assert ws["B6"].value == 3.0             # reported kept; 1.2+1.9=3.1 is 3% off -> no formula
