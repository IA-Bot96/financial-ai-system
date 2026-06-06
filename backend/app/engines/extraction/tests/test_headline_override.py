"""Tests for Bucket-B headline face-truth override."""
import openpyxl

from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable, LineItem, LineItemValue
from app.engines.extraction.services.face_truth import tieout
from app.engines.extraction.services.headline_override import override_headline_metrics


def _company():
    # Audited face truth: total assets 2024=32,873,428 ; 2025=32,988,591.
    return CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.balance_sheet,
                       title="Statement of Financial Position",
                       line_items=[
                           LineItem(label="Total assets", canonical_metric="total_assets",
                                    canonical_category="balance_sheet",
                                    values=[LineItemValue(year=2024, value=32873428.0),
                                            LineItemValue(year=2025, value=32988591.0)]),
                           LineItem(label="Revenue", canonical_metric="revenue",
                                    canonical_category="income_statement",
                                    values=[LineItemValue(year=2024, value=100.0),
                                            LineItemValue(year=2025, value=100.0)]),
                       ])])


def test_failing_headline_cell_is_overridden_with_face_value(tmp_path):
    wb = openpyxl.Workbook()
    bs = wb.active; bs.title = "Balance Sheet"
    bs["A3"], bs["F3"], bs["G3"] = "Particulars", 2024, 2025
    # Wrong computed totals (breakdown leaves understated): far from face truth.
    bs["A28"], bs["F28"], bs["G28"] = "Total assets", "=F14+F26", "=G14+G26"
    bs["F14"], bs["F26"] = 7000000, 22000000     # = 29,000,000 (wrong, face 32,873,428)
    bs["G14"], bs["G26"] = 7000000, 18000000     # = 25,000,000 (wrong, face 32,988,591)
    p = tmp_path / "wb.xlsx"; wb.save(p)

    overrides = override_headline_metrics(p, _company(), tieout, {"Balance Sheet"})
    out = openpyxl.load_workbook(p)["Balance Sheet"]
    assert out["F28"].value == 32873428.0
    assert out["G28"].value == 32988591.0
    assert out["F28"].comment is not None and "OVERRIDE" in out["F28"].comment.text
    assert {(o.coordinate, o.year) for o in overrides} == {("F28", 2024), ("G28", 2025)}


def _pl_company(metric, label, vals):
    return CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label=label, canonical_metric=metric,
                           canonical_category="income_statement",
                           values=[LineItemValue(year=y, value=v) for y, v in vals.items()])])])


def test_expense_overridden_negative_even_when_face_is_positive(tmp_path):
    # Cost of sales face is POSITIVE (a cost reported plain); no formula sibling exists.
    # The P&L additive convention requires it negative -> fallback forces -abs.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "P&L"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A12"], ws["F12"], ws["G12"] = "Cost of sales", 1, 1   # wrong -> overridden, no formula siblings
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = _pl_company("cost_of_sales", "Cost of sales", {2024: 500.0, 2025: 600.0})
    override_headline_metrics(p, company, tieout, {"P&L"})
    out = openpyxl.load_workbook(p)["P&L"]
    assert out["F12"].value == -500.0 and out["G12"].value == -600.0


def test_revenue_overridden_positive(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "P&L"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A5"], ws["F5"], ws["G5"] = "Revenue", 1, 1
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = _pl_company("revenue", "Revenue", {2024: 1000.0, 2025: 1100.0})
    override_headline_metrics(p, company, tieout, {"P&L"})
    out = openpyxl.load_workbook(p)["P&L"]
    assert out["F5"].value == 1000.0 and out["G5"].value == 1100.0


def test_sign_inferred_from_sibling_preserves_loss(tmp_path):
    # profit_after_tax is NOT in the fixed sets (can be a loss). A sibling formula that
    # evaluates negative -> the overridden cell is written negative too (loss preserved).
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "P&L"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A25"], ws["F25"], ws["G25"] = "Profit after taxation", "=-100", 1  # F sibling = -100 (loss)
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = _pl_company("profit_after_tax", "Profit after taxation", {2025: 900.0})
    override_headline_metrics(p, company, tieout, {"P&L"})
    out = openpyxl.load_workbook(p)["P&L"]
    assert out["G25"].value == -900.0          # negative sibling -> loss sign preserved


def test_asset_total_forced_positive_despite_negative_sibling(tmp_path):
    # Regression (#2): a 'Total Non-Current Assets' whose breakdown formula evaluates
    # NEGATIVE (mis-extracted leaves) must NOT be emitted negative — assets use the fixed
    # positive convention, never the sibling-formula sign.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Balance Sheet"
    ws["A3"], ws["F3"], ws["G3"] = "Particulars", 2024, 2025
    ws["A14"], ws["F14"], ws["G14"] = "Total non-current assets", "=-8413955", "=-8014208"
    p = tmp_path / "wb.xlsx"; wb.save(p)
    company = CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.balance_sheet,
                       title="Statement of Financial Position",
                       line_items=[LineItem(label="Total non-current assets",
                           canonical_metric="non_current_assets", canonical_category="balance_sheet",
                           values=[LineItemValue(year=2024, value=8413955.0),
                                   LineItemValue(year=2025, value=8014208.0)])])])
    override_headline_metrics(p, company, tieout, {"Balance Sheet"})
    out = openpyxl.load_workbook(p)["Balance Sheet"]
    assert out["F14"].value == 8413955.0 and out["G14"].value == 8014208.0   # positive, not negated


def test_correct_headline_cell_keeps_its_formula(tmp_path):
    wb = openpyxl.Workbook()
    bs = wb.active; bs.title = "Balance Sheet"
    bs["A3"], bs["F3"], bs["G3"] = "Particulars", 2024, 2025
    bs["A28"], bs["F28"], bs["G28"] = "Total assets", "=F14+F26", "=G14+G26"
    bs["F14"], bs["F26"] = 32873428, 0           # ties out to face exactly
    bs["G14"], bs["G26"] = 30000000, 2988591     # = 32,988,591 ties out
    p = tmp_path / "wb.xlsx"; wb.save(p)

    overrides = override_headline_metrics(p, _company(), tieout, {"Balance Sheet"})
    out = openpyxl.load_workbook(p)["Balance Sheet"]
    assert out["F28"].value == "=F14+F26"        # preserved, not overridden
    assert out["G28"].value == "=G14+G26"
    assert overrides == []


def _income_only_company():
    # Has income-statement face truth (revenue) but NO balance-sheet face truth — the
    # Lucky case: a mis-classified balance sheet yields no primary total_assets table.
    return CompanyResult(company="A", fiscal_years=[2024, 2025], tables=[
        FinancialTable(statement_type=StatementType.income_statement,
                       title="Statement of Profit or Loss",
                       line_items=[LineItem(label="Revenue", canonical_metric="revenue",
                           canonical_category="income_statement",
                           values=[LineItemValue(year=2024, value=100.0),
                                   LineItemValue(year=2025, value=100.0)])])])


def test_coverage_gap_flags_headline_metric_without_face_truth(tmp_path):
    # An emitted headline metric (Total assets) with NO face truth must be flagged as a
    # production-blocking coverage gap (the mis-classified-statement case), while a
    # metric that DOES have face truth (Revenue) is not.
    from app.engines.extraction.services.validation import headline_coverage_gaps
    wb = openpyxl.Workbook()
    bs = wb.active; bs.title = "Balance Sheet"
    bs["A3"], bs["F3"], bs["G3"] = "Particulars", 2024, 2025
    bs["A28"], bs["F28"], bs["G28"] = "Total assets", "=F14+F26", "=G14+G26"  # no face truth
    bs["A6"], bs["F6"], bs["G6"] = "Revenue", 100, 100                         # has face truth
    p = tmp_path / "wb.xlsx"; wb.save(p)

    rows, gaps = headline_coverage_gaps(p, _income_only_company(), {"Balance Sheet"})
    assert gaps == 2                                  # total_assets 2024 + 2025
    assert all(r.status == "NO_FACE_TRUTH" for r in rows)
    assert {(r.metric, r.year) for r in rows} == {("total_assets", 2024), ("total_assets", 2025)}


def test_coverage_gap_exempts_forecast_years(tmp_path):
    # A populated headline cell in a NON-reporting (forecast) year is not a coverage gap.
    from app.engines.extraction.services.validation import headline_coverage_gaps
    wb = openpyxl.Workbook()
    bs = wb.active; bs.title = "Balance Sheet"
    bs["A3"], bs["F3"], bs["G3"], bs["H3"] = "Particulars", 2024, 2025, 2026  # 2026 = forecast
    bs["A28"], bs["F28"], bs["G28"], bs["H28"] = "Total assets", 1, 2, 3
    p = tmp_path / "wb.xlsx"; wb.save(p)
    rows, gaps = headline_coverage_gaps(p, _income_only_company(), {"Balance Sheet"})  # fy 2024,2025
    assert gaps == 2 and all(r.year in (2024, 2025) for r in rows)         # 2026 exempt


def test_breakdown_sheet_not_in_output_set_is_untouched(tmp_path):
    wb = openpyxl.Workbook()
    bs1 = wb.active; bs1.title = "BS1 - Non-Current Assets"
    bs1["A3"], bs1["F3"], bs1["G3"] = "Particulars", 2024, 2025
    bs1["A72"], bs1["F72"], bs1["G72"] = "Total assets", 1, 2   # wrong, but a breakdown note
    p = tmp_path / "wb.xlsx"; wb.save(p)

    # output set excludes the breakdown sheet -> nothing overridden there.
    overrides = override_headline_metrics(p, _company(), tieout, {"Balance Sheet"})
    out = openpyxl.load_workbook(p)["BS1 - Non-Current Assets"]
    assert out["F72"].value == 1 and out["G72"].value == 2
    assert overrides == []
