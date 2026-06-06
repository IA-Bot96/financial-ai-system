"""Tests for delivery/traceability: output-override lineage in the Source Ledger (#5)
and the formula-cache recalc pass (#6)."""
import openpyxl

from app.engines.extraction.pipeline.template_map import MappingPlan
from app.engines.extraction.services.headline_override import Override
from app.engines.extraction.services.validation import recalc_workbook, write_source_ledger


def test_source_ledger_records_output_overrides(tmp_path):
    wb = openpyxl.Workbook(); wb.active.title = "P&L"
    p = tmp_path / "wb.xlsx"; wb.save(p)
    overrides = [Override(sheet="P&L", coordinate="D23", metric="tax_expense", year=2022,
                          was="=-'PL6'!C14", value=-3258105.0, source="millat-2023.pdf p184")]
    write_source_ledger(p, MappingPlan(), overrides)

    led = openpyxl.load_workbook(p)["Source Ledger"]
    rows = [tuple(c.value for c in row) for row in led.iter_rows(min_row=2)]
    assert len(rows) == 1
    sheet, cell, tlabel, mlabel, year, value = rows[0][:6]
    assert (sheet, cell, tlabel, year, value) == ("P&L", "D23", "tax_expense", 2022, -3258105.0)
    assert mlabel == "(headline override)"
    assert "OVERRIDE" in rows[0][-1]


def test_recalc_sets_full_calc_on_load(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"], ws["A2"], ws["A3"] = 1, 2, "=A1+A2"
    p = tmp_path / "wb.xlsx"; wb.save(p)

    materialized = recalc_workbook(p)            # False here (no LibreOffice on PATH)
    out = openpyxl.load_workbook(p)
    assert out.calculation.fullCalcOnLoad is True
    assert out["Sheet"]["A3"].value == "=A1+A2"  # formula preserved
    assert isinstance(materialized, bool)
