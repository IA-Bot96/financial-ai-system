"""Layer 6 — Excel output writer (no-template path) + styled insight sheets.

Generates a workbook with one styled sheet per detected table plus `Insights`
and `Insights Review` sheets, all using the Millat/Lucky template look. The
insight-sheet helper is reusable by the template path too.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl.utils import get_column_letter

from app.core.logging import get_logger
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.financials import FinancialTable
from app.engines.extraction.models.insight import INSIGHT_COLUMNS, Insight
from app.engines.extraction.services import styles as S

logger = get_logger(__name__)

_INVALID_SHEET = re.compile(r"[:\\/?*\[\]]")
_INSIGHT_WIDTHS = [8, 16, 28, 70, 20, 6, 11]  # per INSIGHT_COLUMNS


def _sheet_name(title: str, used: set[str]) -> str:
    name = _INVALID_SHEET.sub(" ", title or "Table").strip()[:31] or "Table"
    base, i = name, 2
    while name.lower() in used:
        suffix = f" {i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def _table_title(table: FinancialTable) -> str:
    base = table.title.strip() if (table.title and table.title.strip()) else \
        table.statement_type.value.replace("_", " ").title()
    # Label the set so consolidated vs unconsolidated tables are distinguishable.
    if table.consolidated is True and "consolidat" not in base.lower():
        base = f"{base} (Consolidated)"
    elif table.consolidated is False and "unconsolidat" not in base.lower() and "separate" not in base.lower():
        base = f"{base} (Unconsolidated)"
    return base


def _write_table(ws, table: FinancialTable) -> None:
    years = sorted(table.years or sorted({v.year for li in table.line_items for v in li.values if v.year}))
    ncols = 1 + len(years)
    last = get_column_letter(ncols)

    # Row 1 — title banner (navy).
    ws.merge_cells(f"A1:{last}1")
    c = ws.cell(1, 1, _table_title(table))
    c.font, c.fill, c.alignment = S.TITLE_FONT, S.TITLE_FILL, S.LEFT

    # Row 2 — unit / currency line (light blue).
    unit_bits = [b for b in (table.currency, table.unit_scale) if b]
    ws.merge_cells(f"A2:{last}2")
    u = ws.cell(2, 1, f"({' in '.join(unit_bits)})" if unit_bits else "")
    u.font, u.fill, u.alignment = S.UNIT_FONT, S.UNIT_FILL, S.LEFT

    # Row 3 — header (navy, white).
    h = ws.cell(3, 1, "Particulars")
    h.font, h.fill, h.alignment = S.HEADER_FONT, S.HEADER_FILL, S.CENTER
    for i, year in enumerate(years, start=2):
        hc = ws.cell(3, i, year)
        hc.font, hc.fill, hc.alignment = S.HEADER_FONT, S.HEADER_FILL, S.CENTER

    # Data rows.
    row = 4
    row_meta: list[dict] = []  # for the subtotal-formula pass
    for li in table.line_items:
        label = (li.label or "").strip()
        if not label:
            continue
        by_year = {v.year: v.value for v in li.values}
        has_value = any(by_year.get(y) is not None for y in years)
        # Prefer the LLM's structural role over the prefix/caps heuristics; fall
        # back to heuristics when no role was provided (older extractions, rules path).
        role = (li.role or "").strip().lower()
        if role in {"leaf", "subtotal", "total", "section_header"}:
            total = role in {"subtotal", "total"}
            section = role == "section_header"
            is_leaf = role == "leaf" and has_value
        else:
            total = S.is_total_row(label)
            section = (not total) and S.is_section_header(label)
            is_leaf = has_value and not total and not section

        a = ws.cell(row, 1, label)
        a.alignment = S.LEFT
        a.font = S.TOTAL_FONT if total else (S.SECTION_FONT if section else S.LABEL_FONT)
        if total:
            a.fill = S.TOTAL_FILL
        elif section:
            a.fill = S.SECTION_FILL

        for i, year in enumerate(years, start=2):
            vc = ws.cell(row, i)
            vc.number_format, vc.alignment = S.NUMBER_FORMAT, S.RIGHT
            value = by_year.get(year)
            if value is not None:
                vc.value = value
            vc.font = S.TOTAL_FONT if total else S.VALUE_FONT
            if total:
                vc.fill = S.TOTAL_FILL
            elif section:
                vc.fill = S.SECTION_FILL

        row_meta.append({
            "row": row,
            "norm": _clabel(label),                    # for component resolution
            "is_total": total,
            "is_section": bool(section),
            "is_leaf": is_leaf,
            "is_contra": bool(li.is_contra),           # subtract this row in its parent
            "components": [_clabel(c) for c in (li.components or [])] or None,
            "by_year": by_year,          # {year: value} as reported
        })
        row += 1

    _apply_formulas(ws, row_meta, years)

    ws.column_dimensions["A"].width = S.LABEL_COL_WIDTH
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = S.VALUE_COL_WIDTH
    ws.freeze_panes = "B4"


def _close(computed: float, reported: float) -> bool:
    """Computed sum agrees with the reported total (within rounding).

    Purely relative (0.1% of the reported magnitude) plus a float-noise epsilon.
    No absolute floor, so a small-magnitude total can't tie out to a materially
    wrong sum: reported 3.0 vs computed 3.1 (3% off) is rejected and the reported
    value is kept. Per-line rounding drift on large totals stays well within
    0.1%; when it doesn't, we simply keep the reported value (a recall miss, never
    a wrong number)."""
    return abs(computed - reported) <= 1e-6 + 1e-3 * abs(reported)


def _clabel(label: str) -> str:
    """Normalize a label for component matching (lowercase, alnum-spaced)."""
    return re.sub(r"[^a-z0-9]+", " ", (label or "").lower()).strip()


def _sign(meta: dict) -> int:
    """+1 normally; -1 for a contra row (deduction/cost stored as a positive)."""
    return -1 if meta.get("is_contra") else 1


def _emit(ws, meta: dict, summands: list[dict], years: list[int], block: bool) -> bool:
    """Write a formula for `meta` summing `summands` (signed), per year column.

    Returns True if at least one year column was formula-ized. Guards keep this
    strictly self-validating (a formula is written only when it provably
    reproduces the report's own stated total — never a wrong number):
      * per-year guard — a column is only formula-ized if >=2 summands carry a
        value that year; otherwise the reported value is kept;
      * tie-out guard — emitted only when the report stated a total for that year
        AND it matches the SIGNED computed sum (contra rows subtracted). A mismatch
        (running subtotal, wrong grouping) keeps the reported value;
      * no-blank-fill — a blank reported total is never back-filled.

    `block=True` lets a contiguous all-positive leaf block render as `SUM(a:b)`;
    otherwise (mixed signs, or an explicit-component/derived total) it renders as
    a signed `a+b-c` reference list so contra terms subtract correctly.
    """
    emitted = False
    for i, year in enumerate(years, start=2):
        col = get_column_letter(i)
        contrib = [s for s in summands if s["by_year"].get(year) is not None]
        if len(contrib) < 2:
            continue  # not enough data this year -> keep reported value
        reported = meta["by_year"].get(year)
        if reported is None:
            continue  # no stated total to validate against -> don't fill a blank cell
        computed = sum(_sign(s) * s["by_year"][year] for s in contrib)
        if not _close(computed, reported):
            continue  # tie-out mismatch -> keep reported value (safe)
        if block and all(_sign(s) == 1 for s in summands):
            rows = [s["row"] for s in summands]   # full leaf block (blanks -> 0 in Excel)
            ref = f"SUM({col}{min(rows)}:{col}{max(rows)})"
        else:                                     # signed reference list: a+b-c
            parts = []
            for j, s in enumerate(contrib):
                cell, pos = f"{col}{s['row']}", _sign(s) == 1
                parts.append((cell if pos else f"-{cell}") if j == 0
                             else (f"+{cell}" if pos else f"-{cell}"))
            ref = "".join(parts)
        ws.cell(meta["row"], i).value = f"={ref}"
        emitted = True
    return emitted


def _resolve_components(comp_norms: list[str], index: dict, total_row: int) -> list[dict] | None:
    """Resolve a total's declared component labels to their rows in this table.

    Returns the summand rows (>=2, de-duplicated) or None if any component can't be
    matched — in which case the caller falls back to positional grouping.
    """
    out: list[dict] = []
    for cn in comp_norms:
        cands = index.get(cn)
        if not cands:
            return None
        above = [m for m in cands if m["row"] < total_row]
        out.append(max(above, key=lambda m: m["row"]) if above
                   else min(cands, key=lambda m: abs(m["row"] - total_row)))
    seen: set[int] = set()
    uniq = [m for m in out if not (m["row"] in seen or seen.add(m["row"]))]
    return uniq if len(uniq) >= 2 else None


def _apply_formulas(ws, row_meta: list[dict], years: list[int]) -> None:
    """Emit semantic formulas for total rows (no-template path), name-independent.

    For each total/subtotal row:
      * if the extractor declared explicit `components`, sum exactly those rows
        (with contra terms subtracted) — this handles running subtotals and nested
        totals that positional grouping can't; otherwise
      * positional fallback: a contiguous leaf block above -> `SUM`; a block of
        subtotals (+ stray leaf) -> signed addition (e.g. TOTAL OPEX = Dist + Admin).

    Every formula still runs through `_emit`, so the per-year / tie-out / no-blank
    guards apply uniformly — a wrong role/component hint can never yield a wrong
    number, only a fallback to the reported value.
    """
    # label -> rows usable as summands (leaves and totals; not section headers).
    index: dict[str, list[dict]] = {}
    for m in row_meta:
        if m["is_leaf"] or m["is_total"]:
            index.setdefault(m["norm"], []).append(m)

    pending: list[dict] = []      # contiguous leaves since the last total/header
    subtotals: list[dict] = []    # subtotal rows available to a higher-level total

    for meta in row_meta:
        if meta["is_section"]:
            pending = []          # a heading bounds the leaf block (no bleed across)
            continue
        if not meta["is_total"]:
            if meta["is_leaf"]:
                pending.append(meta)
            continue

        # Explicit components first (when the extractor declared them and they resolve).
        resolved = _resolve_components(meta["components"], index, meta["row"]) \
            if meta.get("components") else None
        if resolved is not None:
            if _emit(ws, meta, resolved, years, block=False):
                subtotals = [meta]
            pending = []
            continue

        # Positional fallback: contiguous leaves -> SUM; else subtotals(+leaf) -> add.
        if len(pending) >= 2:
            _emit(ws, meta, pending, years, block=True)
            subtotals.append(meta)
        else:
            summands = subtotals + pending
            if len(summands) >= 2 and _emit(ws, meta, summands, years, block=False):
                subtotals = [meta]   # subsumes its components; can feed a higher total
            # if nothing was emitted, keep the accumulated subtotals (no discard)
        pending = []


def write_insights_sheet(ws, insights: list[Insight]) -> None:
    for col, (name, width) in enumerate(zip(INSIGHT_COLUMNS, _INSIGHT_WIDTHS), start=1):
        hc = ws.cell(1, col, name)
        hc.font, hc.fill, hc.alignment = S.HEADER_FONT, S.HEADER_FILL, S.CENTER
        ws.column_dimensions[get_column_letter(col)].width = width
    for r, ins in enumerate(insights, start=2):
        for col, value in enumerate(ins.to_row(), start=1):
            cell = ws.cell(r, col, value)
            cell.font = S.VALUE_FONT
            cell.alignment = S.LEFT if col in (3, 4, 5) else S.CENTER
            if col == len(INSIGHT_COLUMNS):  # Confidence
                cell.number_format = S.CONFIDENCE_FORMAT
    ws.freeze_panes = "A2"


def write_workbook(
    tables: list[FinancialTable],
    insights: list[Insight],
    insights_review: list[Insight],
    output_path: str | Path,
) -> Path:
    """No-template output: one styled sheet per table + Insights / Insights Review."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for table in tables:
        _write_table(wb.create_sheet(_sheet_name(_table_title(table), used)), table)

    write_insights_sheet(wb.create_sheet("Insights"), insights)
    if insights_review:
        write_insights_sheet(wb.create_sheet("Insights Review"), insights_review)

    if not wb.sheetnames:
        wb.create_sheet("Empty")
    wb.save(output_path)
    logger.info("Wrote workbook %s (%d tables, %d insights)", output_path, len(tables), len(insights))
    return Path(output_path)


def write_company_workbook(company: CompanyResult, output_path: str | Path) -> Path:
    return write_workbook(company.tables, company.insights, company.insights_review, output_path)


def append_insights_sheets(
    workbook_path: str | Path,
    insights: list[Insight],
    insights_review: list[Insight],
) -> Path:
    """Add styled Insights / Insights Review sheets to an existing workbook
    (used by the template path after the breakdown sheets are populated)."""
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    for name in ("Insights", "Insights Review"):
        if name in wb.sheetnames:
            del wb[name]
    write_insights_sheet(wb.create_sheet("Insights"), insights)
    if insights_review:
        write_insights_sheet(wb.create_sheet("Insights Review"), insights_review)
    wb.save(workbook_path)
    return Path(workbook_path)
