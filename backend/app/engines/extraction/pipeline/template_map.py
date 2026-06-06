"""Layer 5 — Template mapping (rule-based + label matching).

Populate a template's breakdown input sheets from a multi-year CompanyResult:

  - Detect each sheet's year-header row and which columns are HISTORICAL vs
    FORECAST (forecast columns are never written).
  - Skip computed/output sheets (mostly formulas -> low empty fraction) and
    non-data sheets (no year header).
  - For each leaf input row (mixed-case label, writable/empty value cells), match
    the label to an extracted line item (scoped to the sheet's statement type via
    the classifier) and write the value for each historical year — only into
    empty, non-formula cells, so subtotals/formulas are preserved.

`build_plan` is pure (reads the template, returns a MappingPlan). `apply_plan`
writes the plan into a copy of the template and saves it.
"""
from __future__ import annotations

import re
from pathlib import Path

from app.core.config import get_settings
from app.core.logging import get_logger
from app.engines.extraction.models.common import StatementType
from app.engines.extraction.models.company import CompanyResult
from app.engines.extraction.models.mapping import CellWrite, MappingPlan
from app.engines.extraction.services.classifier import TableClassifier
from app.engines.extraction.services.metric_resolver import get_resolver, spaced

logger = get_logger(__name__)

from app.engines.extraction.services.face_truth import (  # noqa: E402
    CROSS_FAMILY_OK as _CROSS_FAMILY_OK,
    KEY_METRICS as _KEY_METRICS,
    _TYPE_FAMILY,
    tieout as _tieout,
)

# Sheet families where a wrong value is high-impact (equity/solvency). On these,
# a candidate must be CONFIRMED same-polarity — "block when unsure" (#5).
_HIGH_RISK_FAMILIES = {StatementType.share_capital_reserves, StatementType.equity_changes}

# Curated metric -> polarity, for cases where the line's home type is a mixed face
# table (polarity None) and there's no other signal.
_EQUITY_METRICS = frozenset({
    "share_capital", "paid_up_capital", "issued_capital", "reserves", "capital_reserves",
    "revenue_reserves", "general_reserve", "retained_earnings", "unappropriated_profit",
    "equity", "total_equity_and_liabilities", "bonus_shares", "fair_value_reserve",
    "surplus_on_revaluation",
})
_ASSET_METRICS = frozenset({
    "cash_and_cash_equivalents", "cash_and_bank_balances", "cash_in_hand", "trade_debts",
    "trade_receivables", "stock_in_trade", "stores_spares_loose_tools", "property_plant_equipment",
    "operating_fixed_assets", "long_term_investments", "short_term_investments", "intangible_assets",
    "loans_and_advances", "total_assets", "current_assets", "non_current_assets",
})
_LIABILITY_METRICS = frozenset({
    "trade_payables", "short_term_borrowings", "long_term_financing", "lease_liabilities",
    "contract_liabilities", "accrued_liabilities", "total_liabilities", "current_liabilities",
    "non_current_liabilities", "unclaimed_dividend",
})


def _metric_polarity(metric: str | None) -> str | None:
    if not metric:
        return None
    if metric in _EQUITY_METRICS:
        return "equity"
    if metric in _ASSET_METRICS:
        return "asset"
    if metric in _LIABILITY_METRICS:
        return "liability"
    return None


# Sheet titles are explicit in these templates; the classifier often returns None
# for them, so derive polarity from the title as the reliable signal.
_TITLE_POLARITY = (
    ("share capital", "equity"), ("reserves", "equity"), ("equity", "equity"),
    ("liabilit", "liability"),
    ("non-current asset", "asset"), ("current asset", "asset"), ("asset", "asset"),
    ("other income", "income"), ("revenue", "income"),
    ("cost of sales", "expense"), ("finance cost", "expense"),
    ("levy", "expense"), ("taxation", "expense"), ("expense", "expense"),
)


def _sheet_polarity_from_title(title: str) -> str | None:
    t = (title or "").lower()
    for kw, pol in _TITLE_POLARITY:
        if kw in t:
            return pol
    return None

# (label, family, section) -> metric.  Context is part of the key (a label can
# resolve differently per sheet family / section).
_ROW_METRIC_CACHE: dict[tuple[str, object, str], str | None] = {}


def _row_metric(label: str, family: object = None, section: str = "") -> str | None:
    """Resolve a template row label to a canonical metric, SCOPED by the sheet's
    statement type (P2): a metric confidently belonging to a different statement
    family than the sheet is demoted to None. Cached by the full context tuple so
    one sheet's decision can't leak into another."""
    from app.engines.extraction.services.face_truth import metric_incompatible
    key = (label, family, section or "")
    if key not in _ROW_METRIC_CACHE:
        m = get_resolver().resolve(label)
        if m and isinstance(family, StatementType) and \
                metric_incompatible(m.canonical_key, m.category, family):
            _ROW_METRIC_CACHE[key] = None        # scoped-out: wrong family for this sheet
        else:
            _ROW_METRIC_CACHE[key] = m.canonical_key if m else None
    return _ROW_METRIC_CACHE[key]


def _build_face_truth(company: CompanyResult) -> dict[tuple[str, int], float]:
    """{(metric, year): value} from PRIMARY face statements (P1, value only)."""
    from app.engines.extraction.services.face_truth import build_face_truth
    return {k: val for k, (val, _src) in build_face_truth(company.tables).items()}

_YEAR_RE = re.compile(r"(?<!\d)(?:19|20)\d{2}(?!\d)")
# Generic section words shared across sections (don't distinguish Local vs Export).
_SECTION_STOPWORDS = {
    "sales", "cost", "costs", "expense", "expenses", "income", "revenue",
    "total", "net", "gross", "and", "the", "of", "to", "from", "other",
    "note", "notes",
}


# --- statement-type containment (lever 1) ---
# A breakdown sheet's data may be extracted under its parent FACE statement (or a
# sibling breakdown), e.g. "Cash in hand" landing in the whole `balance_sheet`
# table instead of the `current_assets` breakdown. So when a sheet's own type
# yields no match for a row, widen the candidate pool to its statement family.
_FACE_GROUPS: dict[StatementType, set[StatementType]] = {
    StatementType.balance_sheet: {
        StatementType.non_current_assets, StatementType.current_assets,
        StatementType.share_capital_reserves, StatementType.non_current_liabilities,
        StatementType.current_liabilities, StatementType.equity_changes,
    },
    StatementType.income_statement: {
        StatementType.revenue, StatementType.cost_of_sales, StatementType.operating_expenses,
        StatementType.other_income, StatementType.finance_cost, StatementType.taxation,
        StatementType.oci,
    },
}


def _related_types(st: StatementType | None) -> set[StatementType]:
    """The statement family a sheet's type belongs to: itself + its face + the
    face's other parts. Used as a widened (fallback) candidate pool."""
    if st is None:
        return set()
    related: set[StatementType] = {st}
    for face, parts in _FACE_GROUPS.items():
        if st == face or st in parts:
            related.add(face)
            related |= parts
    return related


# --- label normalization for matching (lever 2) ---
# Applied on top of `spaced()` so corporate/financial abbreviations don't sink a
# fuzzy match below threshold (e.g. "(Pvt) Ltd" vs "(Private) Limited").
_ABBREV = {
    "pvt": "private", "ltd": "limited", "corp": "corporation",
    "incl": "including", "amt": "amount", "recd": "received",
    "accum": "accumulated", "invt": "investment", "mfg": "manufacturing",
    "wppf": "workers profit participation fund", "wwf": "workers welfare fund",
}


def _match_norm(label: str) -> str:
    """Normalize a label for fuzzy matching: `spaced()` + abbreviation expansion."""
    return " ".join(_ABBREV.get(tok, tok) for tok in spaced(label).split())


def _significant_tokens(section_norm: str) -> set[str]:
    return {t for t in section_norm.split() if len(t) >= 3 and t not in _SECTION_STOPWORDS}


# --- cell helpers ---

def _is_formula(cell) -> bool:
    if getattr(cell, "data_type", None) == "f":
        return True
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


def _is_empty(cell) -> bool:
    return cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())


def _as_year(value) -> int | None:
    if isinstance(value, (int, float)) and 1990 <= int(value) <= 2100:
        return int(value)
    if isinstance(value, str):
        m = _YEAR_RE.search(value)
        if m:
            return int(m.group())
    return None


def _is_section_header(label: str) -> bool:
    letters = [c for c in label if c.isalpha()]
    if len(letters) < 3:
        return True
    return sum(c.isupper() for c in letters) / len(letters) >= 0.7


# --- sheet structure detection ---

def _detect_year_columns(ws) -> tuple[dict[int, int], int, int | None]:
    """Return (historical {col->year}, header_row, forecast_start_col).

    Forecast boundary is taken from a 'forecast' marker cell if present; columns
    at/after it are excluded from the historical map.
    """
    best: dict[int, int] = {}
    header_row = 1
    scan = min(ws.max_row, 8)
    for r in range(1, scan + 1):
        cols = {c: y for c in range(1, ws.max_column + 1)
                if (y := _as_year(ws.cell(r, c).value)) is not None}
        if len(cols) > len(best):
            best, header_row = cols, r

    forecast_start: int | None = None
    for r in range(1, scan + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "forecast" in v.lower():
                forecast_start = c if forecast_start is None else min(forecast_start, c)
    historical = {c: y for c, y in best.items() if forecast_start is None or c < forecast_start}
    return historical, header_row, forecast_start


def _empty_fraction(ws, year_cols: dict[int, int], header_row: int) -> float:
    total = empty = 0
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str) or not label.strip():
            continue
        for c in year_cols:
            total += 1
            if _is_empty(ws.cell(r, c)) and not _is_formula(ws.cell(r, c)):
                empty += 1
    return (empty / total) if total else 0.0


def _crosssheet_fraction(ws, year_cols: dict[int, int], header_row: int) -> float:
    """Fraction of FORMULA data cells that pull from another sheet (contain '!').

    The assembled output statements (P&L, Balance Sheet) are built almost entirely from
    `='BS1 - …'!`/`='PL1 - …'!` references (~0.6–0.8); breakdown notes use intra-sheet
    sums + leaf values (0.0). So a high fraction marks an OUTPUT statement even when its
    empty-fraction is just over the computed-sheet threshold (the Millat balance sheet)."""
    total = cross = 0
    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str) or not label.strip():
            continue
        for c in year_cols:
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                total += 1
                if "!" in v:
                    cross += 1
    return (cross / total) if total else 0.0


def _sheet_statement_type(ws, header_row: int, classifier: TableClassifier) -> StatementType | None:
    title = ws.cell(1, 1).value or ""
    labels = [
        str(ws.cell(r, 1).value)
        for r in range(header_row + 1, min(ws.max_row, header_row + 25) + 1)
        if isinstance(ws.cell(r, 1).value, str)
    ]
    signature = f"{title} {ws.title} " + " ".join(labels)
    result = classifier.classify(signature)
    return None if result.needs_review else result.statement_type


# --- candidates from extracted data ---

def _build_candidates(company: CompanyResult):
    """Return (by_type, pooled): lists of (label_norm, section_norm, line) per type.

    Templates target the UNCONSOLIDATED set, so when a statement type has both
    sets we use the unconsolidated (or unknown) tables and skip the consolidated
    ones as mapping candidates.
    """
    from collections import defaultdict

    tables_by_type: dict[StatementType, list] = defaultdict(list)
    for table in company.tables:
        tables_by_type[table.statement_type].append(table)

    by_type: dict[StatementType, list[tuple[str, str, object, object]]] = {}
    pooled: list[tuple[str, str, object, object]] = []
    from app.engines.extraction.services.face_truth import table_role_of
    for st, tables in tables_by_type.items():
        # P2: never offer analytical (ratio / six-year-summary / %) lines as mapping
        # candidates — they pollute matches and aren't audited values.
        tables = [t for t in tables if table_role_of(t) != "analytical"]
        preferred = [t for t in tables if t.consolidated is not True]  # False or None
        chosen = preferred or tables  # fall back to consolidated if that's all we have
        bucket: list[tuple[str, str, object, object]] = []
        for table in chosen:
            for li in table.line_items:
                # carry the line's HOME statement type (4th field) for the polarity guard
                entry = (_match_norm(li.label), spaced(li.section or ""), li, table.statement_type)
                bucket.append(entry)
                pooled.append(entry)
        by_type[st] = bucket
    return by_type, pooled


def _looks_total(label: str) -> bool:
    low = (label or "").strip().lower()
    return low.startswith(("total", "gross ", "net ", "subtotal", "operating profit", "profit "))


def _section_overlap(template_section_norm: str, cand_section_norm: str) -> int:
    """Count of shared distinctive section tokens (0 = unrelated/uninformative)."""
    return len(_significant_tokens(template_section_norm) & _significant_tokens(cand_section_norm))


def _best_candidate(template_label: str, template_section: str, candidates, threshold: float,
                    row_metric: str | None = None, sheet_polarity: str | None = None,
                    template_is_total: bool = False, strict_polarity: bool = False,
                    sheet_family: str | None = None):
    """Best extracted line for a template row by LABEL (recall-first), with the
    section used only as a tie-breaker. Conflicts where one extracted line is
    claimed by two template rows are resolved globally in `build_plan`.

    Guards reject a high-label-score-but-wrong candidate:
      * role guard (#7) — a `total`/`subtotal` line is rejected only for a LEAF
        template row (total->leaf); total->total is allowed. Role is inferred from
        the candidate label when the `role` field is absent;
      * metric-agreement (#2) — confident-but-different canonical metric -> reject;
      * polarity guard (#5) — when both the sheet and the candidate have a known
        side (asset/liability/equity/income/expense) that conflicts, reject (e.g.
        a cash/asset line for a share-capital/equity row), unless the metric is a
        known cross-family one (deferred tax, depreciation, …).
    """
    from rapidfuzz import fuzz

    from app.engines.extraction.models.common import polarity as _polarity

    q_label, q_section = _match_norm(template_label), spaced(template_section)
    best, best_label_score, best_combined = None, 0.0, -1.0
    for cand_label, cand_section, line, home_type in candidates:
        cand_role = (getattr(line, "role", None) or "").lower()
        cand_is_total = cand_role in {"total", "subtotal"} or (not cand_role and _looks_total(line.label))
        if cand_is_total and not template_is_total:
            continue  # total->leaf rejected; total->total allowed
        cm = getattr(line, "canonical_metric", None)
        if row_metric and cm and cm != row_metric:
            continue  # different concept (e.g. cash vs share capital) -> reject
        # Statement-family gate: a balance-sheet note row must never take an income-family
        # line (the "Cost of sales" -35M bleed into BS1) and vice-versa. Robust even when
        # polarity can't be inferred. Cross-family metrics (depreciation, taxes-paid, …) and
        # unknown families are exempt — never block on a missing signal.
        if sheet_family and cm not in _CROSS_FAMILY_OK:
            home_family = _TYPE_FAMILY.get(home_type)
            if home_family and home_family != sheet_family:
                continue
        if sheet_polarity and cm not in _CROSS_FAMILY_OK:
            cand_pol = _polarity(home_type, getattr(line, "section", None)) or _metric_polarity(cm)
            if cand_pol and cand_pol != sheet_polarity:
                continue  # cross-statement polarity conflict -> reject
            if strict_polarity and cand_pol != sheet_polarity:
                continue  # high-risk sheet: require a CONFIRMED same-polarity candidate
        label_score = fuzz.token_set_ratio(q_label, cand_label)
        if label_score < threshold:
            continue
        # Prefer a section-matching candidate when labels tie (e.g. Local vs Export).
        combined = label_score + (5 if _section_overlap(q_section, cand_section) else 0)
        if combined > best_combined:
            best, best_label_score, best_combined = line, label_score, combined
    return (best, best_label_score) if best is not None else (None, 0.0)


# --- main ---

def build_plan(company: CompanyResult, template_path: str | Path) -> MappingPlan:
    import openpyxl

    settings = get_settings()
    threshold = settings.template_match_threshold * 100
    classifier = TableClassifier()
    by_type, pooled = _build_candidates(company)
    face = _build_face_truth(company)   # audited primary-face truth (P1)

    wb = openpyxl.load_workbook(template_path, data_only=False)
    plan = MappingPlan()
    try:
        # Pass 1: collect best label matches across ALL sheets (own type first, then
        # the widened family with a section-compatibility gate).
        matches = []  # dicts: ws, row, writable, lbl, template_section, line, score, own_tier
        for ws in wb.worksheets:
            year_cols, header_row, _ = _detect_year_columns(ws)
            if len(year_cols) < 2:
                plan.sheets_skipped.append(ws.title)
                continue
            # An output statement is either mostly-filled-with-formulas (low empty
            # fraction) OR predominantly cross-sheet pulls (a balance sheet that pulls
            # every line from the BS1–BS5 breakdowns but still has empty forecast cells).
            if (_empty_fraction(ws, year_cols, header_row) < settings.template_min_empty_fraction
                    or _crosssheet_fraction(ws, year_cols, header_row) >= 0.30):
                plan.sheets_skipped.append(ws.title)  # computed/output sheet
                plan.formula_sheets.append(ws.title)  # ...and it IS an output statement
                continue

            plan.sheets_processed.append(ws.title)
            st = _sheet_statement_type(ws, header_row, classifier)
            from app.engines.extraction.models.common import polarity as _polarity
            title_pol = _sheet_polarity_from_title(ws.title)
            sheet_polarity = (_polarity(st) if st else None) or title_pol
            strict_polarity = (st in _HIGH_RISK_FAMILIES) or (title_pol == "equity")
            # Statement family of this sheet (income/balance/cash) for the family gate.
            sheet_family = _TYPE_FAMILY.get(st) if st else (
                "balance" if title_pol in ("asset", "liability", "equity") else None)
            # Tier 1: the sheet's own statement type (precision-first).
            own = by_type.get(st) or []
            # Tier 2: widen to the statement family (face + sibling breakdowns) so a
            # line extracted into a parent/sibling table is still reachable; falls
            # back to the full pool only when the family yields nothing.
            widened: list = []
            for rt in _related_types(st):
                widened.extend(by_type.get(rt, []))
            widened = widened or pooled

            template_section = ""
            for r in range(header_row + 1, ws.max_row + 1):
                label = ws.cell(r, 1).value
                if not isinstance(label, str) or not label.strip():
                    continue
                lbl = label.strip()
                if _is_section_header(lbl):
                    template_section = lbl  # e.g. LOCAL SALES / EXPORT SALES
                    continue
                writable = {c: y for c, y in year_cols.items()
                            if _is_empty(ws.cell(r, c)) and not _is_formula(ws.cell(r, c))}
                if not writable or not (own or widened):
                    continue
                # Own type first; only widen to the family when the sheet's own type
                # has no candidate. The global dedup below routes each extracted line
                # to a single best (home-type) sheet, so widened matches can't bleed.
                rm = _row_metric(lbl, st, template_section)
                is_total = _looks_total(lbl)
                line, score = _best_candidate(lbl, template_section, own, threshold, row_metric=rm,
                                              sheet_polarity=sheet_polarity, template_is_total=is_total,
                                              strict_polarity=strict_polarity, sheet_family=sheet_family)
                own_tier = line is not None
                if line is None:
                    line, score = _best_candidate(lbl, template_section, widened, threshold, row_metric=rm,
                                                  sheet_polarity=sheet_polarity, template_is_total=is_total,
                                                  strict_polarity=strict_polarity, sheet_family=sheet_family)
                if line is None:
                    # #4 statement-total fallback: a writable, empty row that resolves
                    # to a HEADLINE metric is filled from the audited face truth (kills
                    # finance-cost / tax / OCI zeros). The key-metric resolution is the
                    # signal it's a statement-level line — labels like "Finance cost" or
                    # "Taxation" need not be prefixed "Total" (M1: key metrics only, so
                    # note subsections — which don't resolve to a key metric — are safe).
                    if rm in _KEY_METRICS:
                        filled = False
                        for col, year in writable.items():
                            tv = face.get((rm, year))
                            if tv is None:
                                continue
                            plan.writes.append(CellWrite(
                                sheet=ws.title, coordinate=ws.cell(r, col).coordinate,
                                year=year, value=tv, template_label=lbl,
                                matched_label=f"[face:{rm}]", confidence=1.0,
                                note="fallback:face_total",
                            ))
                            filled = True
                        if filled:
                            continue
                    plan.unmatched_template_labels.append(lbl)
                    continue
                matches.append({
                    "ws": ws, "row": r, "writable": writable, "lbl": lbl,
                    "template_section": template_section, "line": line,
                    "score": score, "own_tier": own_tier, "row_metric": rm,
                })

        # Pass 2: GLOBAL conflict resolution — one extracted line maps to at most one
        # template row, ACROSS sheets. Prefer the line's home-type sheet (own_tier) so a
        # liability line claimed by an asset sheet via widening loses to its own sheet;
        # then by section overlap, then label score. Losers are recorded as unmatched.
        from collections import defaultdict
        by_line: dict[int, list] = defaultdict(list)
        for m in matches:
            by_line[id(m["line"])].append(m)
        winners = []
        for group in by_line.values():
            winner = max(group, key=lambda m: (
                m["own_tier"],
                _section_overlap(spaced(m["template_section"]), spaced(m["line"].section or "")),
                m["score"],
            ))
            winners.append(winner)
            for m in group:
                if m is not winner:
                    plan.unmatched_template_labels.append(m["lbl"])

        # Write the surviving matches, gated by a statement-level tie-out: a value
        # placed in a headline-total row must reconcile to the audited face
        # statement, or it is WITHHELD (recorded for audit, never written).
        for m in winners:
            line = m["line"]
            # Validate against the metric the ROW is supposed to hold (falling back to
            # the matched line's metric) — so a value landing in the wrong total row is
            # checked against what that row should be, not against the line it matched.
            key_metric = m["row_metric"] or getattr(line, "canonical_metric", None)
            by_year = {v.year: v for v in line.values}
            # #6: NO sign-flipping here. Template writable rows are note/breakdown
            # sheets that keep their own positive-magnitude convention; aligning their
            # sign to the face value (a different convention) would be wrong. A genuine
            # sign conflict instead fails tie-out and is withheld (safe).
            for col, year in m["writable"].items():
                v = by_year.get(year)
                if v is None or v.value is None:
                    continue
                value = v.value
                truth = face.get((key_metric, year)) if key_metric in _KEY_METRICS else None
                cw = CellWrite(
                    sheet=m["ws"].title,
                    coordinate=m["ws"].cell(m["row"], col).coordinate,
                    year=year, value=value, template_label=m["lbl"],
                    matched_label=line.label, confidence=m["score"] / 100.0,
                    source_report_year=v.source_report_year, source=v.source,
                )
                if truth is not None and not _tieout(value, truth):
                    cw.note = f"withheld:tieout({key_metric} face={truth})"
                    plan.withheld.append(cw)
                else:
                    plan.writes.append(cw)
    finally:
        wb.close()

    logger.info(
        "Template plan: %d writes across %d sheet(s); %d skipped, %d unmatched, %d withheld (tie-out)",
        len(plan.writes), len(plan.sheets_processed), len(plan.sheets_skipped),
        len(plan.unmatched_template_labels), len(plan.withheld),
    )
    if plan.withheld:
        for cw in plan.withheld[:10]:
            logger.warning("  WITHHELD %s!%s %s=%s (%s) — %s",
                           cw.sheet, cw.coordinate, cw.template_label, cw.value,
                           cw.matched_label, cw.note)
    return plan


def apply_plan(plan: MappingPlan, template_path: str | Path, output_path: str | Path,
               company: str | None = None) -> Path:
    """Write the plan into a copy of the template (formulas/forecast untouched).

    Also replaces static "Source: …" label cells with a DYNAMIC reference naming
    the report years actually used on that sheet (traceability #9 / MIL-OCR-014)."""
    import openpyxl

    # report years actually used per sheet (from each write's source).
    years_by_sheet: dict[str, set[int]] = {}
    for w in plan.writes:
        ry = w.source_report_year
        if ry:
            years_by_sheet.setdefault(w.sheet, set()).add(ry)

    wb = openpyxl.load_workbook(template_path, data_only=False)
    try:
        for w in plan.writes:
            wb[w.sheet][w.coordinate] = w.value
        # rewrite static "Source:" labels with the real per-sheet source years.
        co = company or "the company"
        for ws in wb.worksheets:
            yrs = sorted(years_by_sheet.get(ws.title, []))
            if not yrs:
                continue
            label = f"Source: {co} Annual Reports {yrs[0]}-{yrs[-1]}" if len(yrs) > 1 \
                else f"Source: {co} Annual Report {yrs[0]}"
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.lower().startswith("source:"):
                        c.value = label
        wb.save(output_path)
    finally:
        wb.close()
    logger.info("Wrote %d values into %s", len(plan.writes), output_path)
    return Path(output_path)
