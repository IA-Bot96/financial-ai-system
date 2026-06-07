"""Generate a small, deterministic .xlsx that reproduces every round-trip-contract
feature the extraction pipeline emits — so the frontend editor's round-trip test has a
committable fixture that doesn't depend on ephemeral `storage/sessions/` output.

It deliberately uses openpyxl (the same writer the pipeline uses), so the comment part
layout is the REAL `xl/comments/comment1.xml` + `xl/drawings/commentsDrawingN.vml` that
broke the editor's `stripComments` regex — not a synthetic shape.

    python -m scripts.make_roundtrip_fixture [out_path]

Default out: storage/fixtures/roundtrip_fixture.xlsx . Copy it into the frontend's
scripts/fixtures/ (or point the test at it via FIXTURE_XLSX=...) and commit it.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

NAVY = "1F4E79"
MONEY = r"#,##0;(#,##0);\-"   # thousands, negatives in parens, dash for zero (pipeline's format)


def build(out_path: Path) -> Path:
    wb = openpyxl.Workbook()

    # --- Sheet 1: a styled statement with a formula, a merge, a frozen pane, a comment ---
    bs = wb.active
    bs.title = "BS"
    bs.merge_cells("A1:C1")                                  # title banner (merged)
    t = bs["A1"]
    t.value = "Balance Sheet"
    t.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left")

    for col, yr in (("B", 2023), ("C", 2024)):              # header row (row 3)
        h = bs[f"{col}3"]
        h.value = yr
        h.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=NAVY)
        h.alignment = Alignment(horizontal="center")
    bs["A3"] = "Particulars"

    bs["A4"], bs["A5"], bs["A6"] = "Cash", "Receivables", "Total current assets"
    bs["B4"] = None                                          # EMPTY input cell -> the edit target
    bs["C4"] = 1500
    bs["B5"], bs["C5"] = 800, 900
    bs["B6"] = "=SUM(B4:B5)"                                 # formula that must survive untouched
    bs["C6"] = "=SUM(C4:C5)"
    for coord in ("B4", "B5", "B6", "C4", "C5", "C6"):
        bs[coord].number_format = MONEY                      # number format must survive
        bs[coord].alignment = Alignment(horizontal="right")
    # A 'validation'-author comment -> openpyxl writes comments/comment1.xml + a VML drawing.
    bs["C6"].comment = Comment("OVERRIDE: substituted audited Total = 2,400. Source: 2024.pdf p12", "validation")
    bs.freeze_panes = "B4"                                   # frozen pane
    bs.column_dimensions["A"].width = 30
    bs.column_dimensions["B"].width = 16
    bs.column_dimensions["C"].width = 16

    # --- Sheet 2: an untouched sheet that must round-trip byte-identical ---
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Scope: P&L + BS only."
    notes["A2"] = "=BS!B6"                                   # cross-sheet formula pull

    wb.calculation.fullCalcOnLoad = True                     # calcPr must be preserved

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _audit(path: Path) -> None:
    import zipfile
    wb = openpyxl.load_workbook(path, data_only=False)
    formulas = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
                   if isinstance(c.value, str) and c.value.startswith("="))
    styled = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.has_style)
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
    comment_parts = [n for n in names if "comment" in n.lower()]
    print(f"  sheets={wb.sheetnames}")
    print(f"  formulas={formulas}  styled_cells={styled}  fullCalcOnLoad={wb.calculation.fullCalcOnLoad}")
    print(f"  comment/vml parts={comment_parts}")


def main() -> None:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("storage/fixtures/roundtrip_fixture.xlsx")
    p = build(out)
    print(f"Wrote {p.resolve()}")
    _audit(p)


if __name__ == "__main__":
    main()
